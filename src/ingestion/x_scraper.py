"""Resumable X ingestion pipeline for the media sentiment project.

Credentials are loaded from environment variables and are never stored in source.
"""


# -----------------------------------------------------------------------------
# Notebook cell 2
# -----------------------------------------------------------------------------
import os
import sys
import re
import csv
import json
import time
import random
import sqlite3
import hashlib
import logging
import threading
import urllib.parse
import html
import shutil
from datetime import datetime, timedelta, timezone
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

PROJECT_ROOT = Path(__file__).resolve().parents[2]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import config_loader, query_registry_loader

import pandas as pd
from colorama import init
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, WebDriverException
from IPython.display import display, HTML
from dotenv import load_dotenv

load_dotenv()
init(autoreset=True)

PIPELINE_CONFIG = config_loader.load_config()
X_CONFIG = PIPELINE_CONFIG.x

if "x" not in PIPELINE_CONFIG.platforms:
    raise ValueError(
        "X collection is disabled: add 'x' to config.yaml platforms."
    )

if not X_CONFIG:
    raise ValueError(
        "X configuration is missing from config.yaml."
    )

PROJECT_CONFIG = {
    "project_start": PIPELINE_CONFIG.date_range.start.isoformat(),
    "project_end": PIPELINE_CONFIG.date_range.end.date().isoformat(),
    "query_version": X_CONFIG["query_version"],
    "collector_version": X_CONFIG["collector_version"],
    "platform": "x",
    "sort_mode": X_CONFIG["sort_mode"],
}

# Collection tuning: favor many shallow date slices over deep scrolling.
SLICE_SPLIT_CONFIG = X_CONFIG["slice_split"]
NAVIGATION_CONFIG = X_CONFIG["navigation"]

MAX_WORKERS = int(X_CONFIG["max_workers"])
MAX_SCROLLS_PER_SLICE = int(X_CONFIG["max_scrolls_per_slice"])
MAX_SCROLLS_MIN_DAY = int(X_CONFIG["max_scrolls_min_day"])
NO_NEW_SCROLL_LIMIT = int(X_CONFIG["no_new_scroll_limit"])

SPLIT_LAST_SCROLL_NEW = int(
    SLICE_SPLIT_CONFIG["last_scroll_new_threshold"]
)
SPLIT_TOTAL_SEEN = int(
    SLICE_SPLIT_CONFIG["total_seen_threshold"]
)
MIN_SLICE_DAYS = int(
    SLICE_SPLIT_CONFIG["minimum_slice_days"]
)

PAGE_LOAD_TIMEOUT = int(
    NAVIGATION_CONFIG["page_load_timeout_seconds"]
)
SAFE_GET_READY_TIMEOUT = int(
    NAVIGATION_CONFIG["ready_timeout_seconds"]
)
STARTUP_STAGGER_SECONDS = float(
    NAVIGATION_CONFIG["startup_stagger_seconds"]
)
STARTUP_MAX_ATTEMPTS = int(
    NAVIGATION_CONFIG["startup_max_attempts"]
)
STARTUP_RETRY_SECONDS = int(
    NAVIGATION_CONFIG["startup_retry_seconds"]
)
SCROLL_MIN_DELAY = float(
    NAVIGATION_CONFIG["scroll_min_delay_seconds"]
)
SCROLL_MAX_DELAY = float(
    NAVIGATION_CONFIG["scroll_max_delay_seconds"]
)
JOB_BACKOFF_SECONDS = int(
    NAVIGATION_CONFIG["job_backoff_seconds"]
)
MAX_JOB_ATTEMPTS = int(
    NAVIGATION_CONFIG["max_job_attempts"]
)

# Optional legacy author-identity backfill. This is NOT run automatically.
BACKFILL_CONFIG = X_CONFIG["backfill"]

BACKFILL_MIN_DELAY = float(
    BACKFILL_CONFIG["min_delay_seconds"]
)
BACKFILL_MAX_DELAY = float(
    BACKFILL_CONFIG["max_delay_seconds"]
)
BACKFILL_READY_TIMEOUT = int(
    BACKFILL_CONFIG["ready_timeout_seconds"]
)

# Progress / durability tuning.
PERSISTENCE_CONFIG = X_CONFIG["persistence"]

RAW_OPERATIONAL_TARGET = int(
    PERSISTENCE_CONFIG["operational_target"]
)
DASHBOARD_REFRESH_SECONDS = float(
    PERSISTENCE_CONFIG["dashboard_refresh_seconds"]
)
BACKUP_INTERVAL_SECONDS = int(
    PERSISTENCE_CONFIG["backup_interval_seconds"]
)
RECOVERY_FSYNC_EVERY = int(
    PERSISTENCE_CONFIG["recovery_fsync_every"]
)


RUNTIME_CONFIG = X_CONFIG["runtime"]
OUTPUT_ROOT_ENV_VAR = RUNTIME_CONFIG["output_root_env_var"]

try:
    from google.colab import drive

    drive.mount("/content/drive", force_remount=False)
    default_output_root = RUNTIME_CONFIG["default_colab_output_root"]
except ImportError:
    default_output_root = RUNTIME_CONFIG["default_local_output_root"]

configured_output_root = os.environ.get(
    OUTPUT_ROOT_ENV_VAR,
    default_output_root,
).strip()

if not configured_output_root:
    raise ValueError(
        f"{OUTPUT_ROOT_ENV_VAR} must not be empty."
    )

DRIVE_DIR = Path(configured_output_root).expanduser()
DRIVE_DIR.mkdir(parents=True, exist_ok=True)
DB_FILE = DRIVE_DIR / 'twitter_data_v4.db'
EXPORT_DIR = DRIVE_DIR / 'exports'
RECOVERY_DIR = DRIVE_DIR / 'recovery_journals'
BACKUP_DIR = DRIVE_DIR / 'backups'
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
RECOVERY_DIR.mkdir(parents=True, exist_ok=True)
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = DRIVE_DIR / 'collector_v4_5.log'
BACKUP_FILE = BACKUP_DIR / 'twitter_data_v4_latest_backup.db'

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(threadName)s | %(message)s',
    handlers=[logging.FileHandler(LOG_FILE, encoding='utf-8'), logging.StreamHandler()],
    force=True,
)

# Required by Raw Schema. Keep one fixed secret salt for the entire project.
# If no environment variable is supplied, create/reuse a private salt file in the project folder.
AUTHOR_SALT_ENV_VAR = RUNTIME_CONFIG["author_salt_env_var"]
SALT_FILENAME = RUNTIME_CONFIG["salt_filename"]
SALT_FILE = DRIVE_DIR / SALT_FILENAME

configured_author_salt = os.environ.get(
    AUTHOR_SALT_ENV_VAR,
    "",
).strip()

if configured_author_salt:
    AUTHOR_SALT = configured_author_salt
elif SALT_FILE.exists():
    AUTHOR_SALT = SALT_FILE.read_text(
        encoding="utf-8"
    ).strip()
else:
    import secrets

    AUTHOR_SALT = secrets.token_hex(32)
    SALT_FILE.write_text(
        AUTHOR_SALT,
        encoding="utf-8",
    )
    try:
        os.chmod(SALT_FILE, 0o600)
    except OSError:
        pass

    print(f"Created persistent project salt at: {SALT_FILE}")

if not AUTHOR_SALT:
    raise ValueError(
        "Author salt is empty. Provide a non-empty "
        f"{AUTHOR_SALT_ENV_VAR} or a valid {SALT_FILE}."
    )

os.environ[AUTHOR_SALT_ENV_VAR] = AUTHOR_SALT

STOP_EVENT = threading.Event()
DB_LOCK = threading.RLock()
STATE_LOCK = threading.RLock()
JOURNAL_LOCKS = {}
SESSION_STARTED_MONO = None
SESSION_STARTED_UTC = None
LAST_BACKUP_MONO = 0.0

print(f"Persistent DB: {DB_FILE}")
print(f"Recovery journals: {RECOVERY_DIR}")
print(f"Exports: {EXPORT_DIR}")
print("Persistence mode: Google Drive write-through + recovery journals + periodic DB backup")

# -----------------------------------------------------------------------------
# Notebook cell 3
# -----------------------------------------------------------------------------
# GITHUB-SAFE ACCOUNT CONFIGURATION
#
# No X/Twitter cookies, auth tokens, proxy addresses, usernames, or passwords are
# stored in source code. Supply them through the environment variable configured
# by x.runtime.accounts_env_var in config.yaml.
# Local developers may keep the variable in a private .env file; `.env` is ignored
# by Git. In Colab, set the variable in the runtime environment before this cell.

ACCOUNTS_ENV_VAR = RUNTIME_CONFIG["accounts_env_var"]


def load_accounts_from_env(env_var=ACCOUNTS_ENV_VAR):
    """Load X account configuration from a private environment variable.

    Parameters
    ----------
    env_var : str
        Environment-variable name containing the account JSON payload.

    Returns
    -------
    list of dict
        Validated runtime account configurations, or an empty list when unset.

    Raises
    ------
    ValueError
        If the environment variable contains invalid account JSON.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    raw = os.environ.get(env_var, '').strip()
    if not raw:
        logging.warning(
            '%s is not set. The notebook can be inspected/exported, but collection '
            'cannot start until account configuration is supplied.',
            env_var,
        )
        return []

    try:
        accounts = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f'{env_var} must contain valid JSON.') from exc

    if not isinstance(accounts, list):
        raise ValueError(f'{env_var} must be a JSON list of account objects.')

    sanitized = []
    for index, item in enumerate(accounts, start=1):
        if not isinstance(item, dict):
            raise ValueError(f'Account #{index} must be a JSON object.')

        cookies = item.get('cookies') or []
        if not isinstance(cookies, list):
            raise ValueError(f'Account #{index} cookies must be a JSON list.')

        sanitized.append({
            'name': str(item.get('name') or f'Account {index}'),
            'proxy_host': str(item.get('proxy_host') or '').strip(),
            'proxy_port': str(item.get('proxy_port') or '').strip(),
            'proxy_user': str(item.get('proxy_user') or ''),
            'proxy_pass': str(item.get('proxy_pass') or ''),
            'cookies': cookies,
        })

    return sanitized


ACCOUNTS = load_accounts_from_env()
print(f'Runtime account configurations loaded: {len(ACCOUNTS)}')

# -----------------------------------------------------------------------------
# Notebook cell 4
# -----------------------------------------------------------------------------
# Query Registry v3 — X only. The YAML registry is the single source of truth.
# XQ-H03 remains inactive because no executed Persian hashtag query was recorded.
X_QUERY_ENTRIES = query_registry_loader.load_all_x_queries()
X_REGISTRY_VERSION = query_registry_loader.get_x_registry_version()

if X_REGISTRY_VERSION != PROJECT_CONFIG["query_version"]:
    raise ValueError(
        "X query-version mismatch: "
        f"config.yaml={PROJECT_CONFIG['query_version']!r}, "
        f"query_registry.yaml={X_REGISTRY_VERSION!r}."
    )

X_QUERIES = [
    {
        "query_id": entry.query_id,
        "family": entry.family,
        "lang": entry.language,
        "logical_query": entry.query_text,
        "risk": entry.risk,
        "entity_anchor": entry.entity_anchor,
        "route": entry.discovery_route,
    }
    for entry in X_QUERY_ENTRIES
]

loaded_query_ids = [entry.query_id for entry in X_QUERY_ENTRIES]
duplicate_query_ids = sorted({
    query_id
    for query_id in loaded_query_ids
    if loaded_query_ids.count(query_id) > 1
})

if not X_QUERY_ENTRIES:
    raise ValueError("No X queries were loaded from query_registry.yaml.")
if duplicate_query_ids:
    raise ValueError(f"Duplicate X query IDs: {duplicate_query_ids}")
if any(entry.platform != "x" for entry in X_QUERY_ENTRIES):
    raise ValueError("Every x_queries entry must have platform='x'.")
if any(entry.language not in {"en", "fa"} for entry in X_QUERY_ENTRIES):
    raise ValueError("Every X query language must be 'en' or 'fa'.")
if any(
    entry.discovery_route not in {"query_search", "hashtag"}
    for entry in X_QUERY_ENTRIES
):
    raise ValueError(
        "Every X discovery_route must be 'query_search' or 'hashtag'."
    )
if any(
    entry.active_from != PIPELINE_CONFIG.date_range.start
    or entry.active_to != PIPELINE_CONFIG.date_range.end.date()
    for entry in X_QUERY_ENTRIES
):
    raise ValueError(
        "Every active X query must use the configured project date range."
    )

QUERY_BY_ID = {q['query_id']: q for q in X_QUERIES}

for q in X_QUERIES:
    if q['risk'] == 'high' and not q['entity_anchor']:
        raise ValueError(f"High-risk query missing entity anchor: {q['query_id']}")

PROJECT_START = datetime.fromisoformat(PROJECT_CONFIG['project_start']).replace(tzinfo=timezone.utc)
PROJECT_END_EXCLUSIVE = (datetime.fromisoformat(PROJECT_CONFIG['project_end']) + timedelta(days=1)).replace(tzinfo=timezone.utc)

def build_project_weeks():
    """Build the ordered project-week calendar for the configured study window.

    Returns
    -------
    list of dict
        Ordered project-week definitions.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    weeks = []
    cursor = PROJECT_START
    idx = 1
    while cursor < PROJECT_END_EXCLUSIVE:
        end = min(cursor + timedelta(days=7), PROJECT_END_EXCLUSIVE)
        weeks.append({
            'project_week': f'W{idx:02d}',
            'start': cursor,
            'end_exclusive': end,
            'is_partial_week': end - cursor < timedelta(days=7),
        })
        cursor = end
        idx += 1
    return weeks

PROJECT_WEEKS = build_project_weeks()
if not os.getenv("SCRAPER_CUSTOM_TOPIC"):
    assert len(PROJECT_WEEKS) == 21
    assert PROJECT_WEEKS[-1]['project_week'] == 'W21'
    assert (PROJECT_WEEKS[-1]['end_exclusive'] - PROJECT_WEEKS[-1]['start']).days == 5
print(f"Loaded {len(X_QUERIES)} X queries × {len(PROJECT_WEEKS)} weeks = {len(X_QUERIES)*len(PROJECT_WEEKS)} initial query-week jobs")

# -----------------------------------------------------------------------------
# Notebook cell 5
# -----------------------------------------------------------------------------
def utc_now():
    """Return the current UTC timestamp in project string format.

    Returns
    -------
    str
        Current UTC timestamp formatted for storage.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    return datetime.now(timezone.utc).isoformat(timespec='seconds').replace('+00:00', 'Z')

def connect_db():
    # DB lives on Google Drive. Serialize writers with DB_LOCK and favor durability over raw write speed.
    """Open a configured SQLite connection to the persistent project database.

    Returns
    -------
    sqlite3.Connection
        Configured SQLite connection.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = sqlite3.connect(DB_FILE, timeout=90, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute('PRAGMA journal_mode=DELETE;')
    except Exception:
        pass
    conn.execute('PRAGMA synchronous=FULL;')
    conn.execute('PRAGMA busy_timeout=90000;')
    conn.execute('PRAGMA foreign_keys=ON;')
    return conn

def init_database():
    """Create the scraper database tables and indexes when they do not yet exist.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    cur = conn.cursor()

    cur.execute("""CREATE TABLE IF NOT EXISTS tweets_raw (
        platform TEXT NOT NULL,
        platform_content_id TEXT NOT NULL,
        content_type TEXT NOT NULL,
        created_at_utc TEXT,
        collected_at_utc TEXT NOT NULL,
        text_raw TEXT,
        author_username TEXT,
        author_display_name TEXT,
        tweet_url TEXT,
        author_hash TEXT NOT NULL,
        project_week TEXT NOT NULL,
        in_window INTEGER NOT NULL,
        is_partial_week INTEGER NOT NULL,
        query_id TEXT NOT NULL,
        collection_run_id TEXT NOT NULL,
        source_id TEXT,
        source_container TEXT,
        source_container_id TEXT,
        source_parent_id TEXT,
        source_parent_title TEXT,
        parent_id TEXT,
        query_version TEXT,
        discovery_route TEXT,
        matched_query_ids TEXT,
        collector_version TEXT,
        permalink_hash TEXT,
        source_total_available INTEGER,
        sampling_applied INTEGER,
        items_kept INTEGER,
        random_seed TEXT,
        engagement_score INTEGER,
        engagement_replies INTEGER,
        engagement_shares INTEGER,
        engagement_quotes INTEGER,
        engagement_views INTEGER,
        engagement_collected_at_utc TEXT,
        author_is_verified INTEGER,
        author_follower_count INTEGER,
        author_account_age_days INTEGER,
        automation_risk_score REAL,
        language_reported TEXT,
        language_detected TEXT,
        language_confidence REAL,
        content_status TEXT,
        geo_method TEXT,
        country_or_region TEXT,
        geo_confidence TEXT,
        geo_granularity TEXT,
        geo_limitations TEXT,
        author_hash_method TEXT,
        content_type_confidence TEXT,
        PRIMARY KEY (platform, platform_content_id)
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS tweet_matches (
        platform TEXT NOT NULL,
        platform_content_id TEXT NOT NULL,
        query_id TEXT NOT NULL,
        project_week TEXT NOT NULL,
        collection_run_id TEXT NOT NULL,
        matched_at_utc TEXT NOT NULL,
        PRIMARY KEY (platform, platform_content_id, query_id, project_week, collection_run_id)
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS x_jobs (
        job_id TEXT PRIMARY KEY,
        query_id TEXT NOT NULL,
        project_week TEXT NOT NULL,
        slice_start TEXT NOT NULL,
        slice_end_exclusive TEXT NOT NULL,
        depth INTEGER NOT NULL DEFAULT 0,
        status TEXT NOT NULL DEFAULT 'pending',
        attempt_count INTEGER NOT NULL DEFAULT 0,
        claimed_by TEXT,
        claimed_at_utc TEXT,
        finished_at_utc TEXT,
        parent_job_id TEXT,
        last_error TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS x_subruns (
        collection_run_id TEXT PRIMARY KEY,
        job_id TEXT NOT NULL,
        worker_name TEXT,
        account_name TEXT,
        platform TEXT,
        query_id TEXT,
        query_text TEXT,
        query_version TEXT,
        project_week TEXT,
        discovery_route TEXT,
        source_id TEXT,
        sort_mode TEXT,
        slice_start TEXT,
        slice_end_exclusive TEXT,
        depth INTEGER,
        started_at_utc TEXT,
        finished_at_utc TEXT,
        returned_count INTEGER,
        stored_count INTEGER,
        records_in_window INTEGER,
        oldest_record_utc TEXT,
        newest_record_utc TEXT,
        error_count INTEGER,
        last_scroll_new INTEGER,
        scrolls_used INTEGER,
        split_triggered INTEGER,
        notes TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS recovery_offsets (
        journal_file TEXT PRIMARY KEY,
        byte_offset INTEGER NOT NULL DEFAULT 0,
        updated_at_utc TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS worker_heartbeat (
        worker_name TEXT PRIMARY KEY,
        account_name TEXT,
        state TEXT,
        query_id TEXT,
        project_week TEXT,
        slice_start TEXT,
        slice_end_exclusive TEXT,
        current_seen INTEGER DEFAULT 0,
        current_stored INTEGER DEFAULT 0,
        session_stored INTEGER DEFAULT 0,
        session_jobs_done INTEGER DEFAULT 0,
        session_errors INTEGER DEFAULT 0,
        started_at_utc TEXT,
        last_heartbeat_utc TEXT,
        note TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS collector_sessions (
        session_id TEXT PRIMARY KEY,
        started_at_utc TEXT,
        finished_at_utc TEXT,
        status TEXT,
        workers INTEGER,
        start_raw_count INTEGER,
        end_raw_count INTEGER,
        notes TEXT
    )""")

    migrate_database_v45(conn)
    conn.commit()
    conn.close()


def migrate_database_v45(conn):
    """Apply the idempotent v4.5 schema migration to an existing database.

    Parameters
    ----------
    conn : sqlite3.Connection
        Open SQLite connection used for the operation.

    Raises
    ------
    sqlite3.Error
        If the schema migration cannot be committed.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    existing = {r['name'] for r in conn.execute("PRAGMA table_info(tweets_raw)").fetchall()}
    additions = {
        'author_username': 'TEXT',
        'author_display_name': 'TEXT',
        'tweet_url': 'TEXT',
    }
    added = []
    for col, decl in additions.items():
        if col not in existing:
            conn.execute(f"ALTER TABLE tweets_raw ADD COLUMN {col} {decl}")
            added.append(col)

    # A direct status URL can be reconstructed from the stable tweet ID even when the
    # original username was intentionally not stored by older versions.
    conn.execute("""
        UPDATE tweets_raw
        SET tweet_url = 'https://x.com/i/web/status/' || platform_content_id
        WHERE (tweet_url IS NULL OR TRIM(tweet_url)='')
          AND platform='x'
          AND platform_content_id IS NOT NULL
          AND TRIM(platform_content_id) <> ''
    """)
    conn.commit()
    if added:
        print("v4.5 schema migration added:", ", ".join(added))

def job_id(query_id, project_week, start_dt, end_dt, depth):
    """Build a deterministic identifier for a query-week slice job.

    Parameters
    ----------
    query_id : str
        Project query identifier.
    project_week : str
        Project-week label, for example ``W01``.
    start_dt : datetime or str
        Inclusive start of the job slice.
    end_dt : datetime or str
        Exclusive end of the job slice.
    depth : int
        Adaptive-splitting depth of the job.

    Returns
    -------
    str
        Deterministic job identifier.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    return f"{query_id}|{project_week}|{start_dt.date().isoformat()}|{end_dt.date().isoformat()}|D{depth}"

def reset_stale_running_jobs():
    """Return stale running jobs to pending state after an interrupted session.

    Returns
    -------
    object
        Result produced by the helper operation.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    with DB_LOCK:
        n = conn.execute("""UPDATE x_jobs
            SET status='pending', claimed_by=NULL, claimed_at_utc=NULL, finished_at_utc=NULL,
                last_error=CASE WHEN last_error IS NULL OR last_error='' THEN 'recovered_after_interruption' ELSE last_error END
            WHERE status='running'""").rowcount
        conn.commit()
    conn.close()
    if n:
        print(f"Recovered {n} stale running job(s) → pending.")
    return n

def seed_initial_jobs(reset_stale_running=True):
    """Seed the initial Query × Project-Week jobs without duplicating existing jobs.

    Parameters
    ----------
    reset_stale_running : bool
        Whether stale running jobs are reset before seeding.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    cur = conn.cursor()
    if reset_stale_running:
        with DB_LOCK:
            cur.execute("""UPDATE x_jobs
                SET status='pending', claimed_by=NULL, claimed_at_utc=NULL, finished_at_utc=NULL
                WHERE status='running'""")
    inserted = 0
    for q in X_QUERIES:
        for w in PROJECT_WEEKS:
            jid = job_id(q['query_id'], w['project_week'], w['start'], w['end_exclusive'], 0)
            cur.execute("""INSERT OR IGNORE INTO x_jobs
                (job_id, query_id, project_week, slice_start, slice_end_exclusive, depth, status)
                VALUES (?, ?, ?, ?, ?, 0, 'pending')""",
                (jid, q['query_id'], w['project_week'], w['start'].date().isoformat(), w['end_exclusive'].date().isoformat()))
            inserted += cur.rowcount
    conn.commit()
    conn.close()
    print(f"Seeded {inserted} new jobs. Existing completed/split jobs were preserved.")

def claim_next_job(conn, worker_name):
    """Atomically claim the next pending collection job for a worker."""
    with DB_LOCK:
        conn.execute('BEGIN IMMEDIATE')

        active_qids = list(QUERY_BY_ID.keys())
        if not active_qids:
            conn.commit()
            return None

        placeholders = ','.join('?' for _ in active_qids)
        sql = f"""SELECT * FROM x_jobs
            WHERE status='pending' AND attempt_count < ? AND query_id IN ({placeholders})
            ORDER BY depth ASC, project_week ASC, query_id ASC, slice_start ASC
            LIMIT 1"""

        params = [MAX_JOB_ATTEMPTS] + active_qids
        row = conn.execute(sql, params).fetchone()

        if not row:
            conn.commit()
            return None

        conn.execute("""UPDATE x_jobs
            SET status='running', claimed_by=?, claimed_at_utc=?, finished_at_utc=NULL,
                attempt_count=attempt_count+1
            WHERE job_id=? AND status='pending'""", (worker_name, utc_now(), row['job_id']))
        conn.commit()
        return conn.execute("SELECT * FROM x_jobs WHERE job_id=?", (row['job_id'],)).fetchone()

def set_job_status(conn, jid, status, error=None):
    """Update the persisted status and optional error message for a job.

    Parameters
    ----------
    conn : sqlite3.Connection
        Open SQLite connection used for the operation.
    jid : str
        Persisted job identifier.
    status : str
        New persisted job status.
    error : str or None
        Optional error text associated with the status change.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    with DB_LOCK:
        if status == 'pending':
            conn.execute("""UPDATE x_jobs SET status=?, finished_at_utc=NULL, last_error=?,
                          claimed_by=NULL, claimed_at_utc=NULL WHERE job_id=?""",
                         (status, error, jid))
        else:
            conn.execute("UPDATE x_jobs SET status=?, finished_at_utc=?, last_error=? WHERE job_id=?",
                         (status, utc_now(), error, jid))
        conn.commit()

def enqueue_split_children(conn, row):
    """Split a dense job into child date ranges and enqueue them transactionally.

    Parameters
    ----------
    conn : sqlite3.Connection
        Open SQLite connection used for the operation.
    row : sqlite3.Row or Mapping
        Persisted job row to process.

    Returns
    -------
    bool
        ``True`` when child jobs are created; otherwise ``False``.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    start = datetime.fromisoformat(row['slice_start']).replace(tzinfo=timezone.utc)
    end = datetime.fromisoformat(row['slice_end_exclusive']).replace(tzinfo=timezone.utc)
    days = (end - start).days
    if days <= MIN_SLICE_DAYS:
        return []
    left_days = max(1, days // 2)
    mid = start + timedelta(days=left_days)
    if mid >= end:
        return []
    children = []
    for a, b in [(start, mid), (mid, end)]:
        jid = job_id(row['query_id'], row['project_week'], a, b, row['depth'] + 1)
        with DB_LOCK:
            conn.execute("""INSERT OR IGNORE INTO x_jobs
                (job_id, query_id, project_week, slice_start, slice_end_exclusive, depth, status, parent_job_id)
                VALUES (?, ?, ?, ?, ?, ?, 'pending', ?)""",
                (jid, row['query_id'], row['project_week'], a.date().isoformat(), b.date().isoformat(),
                 row['depth']+1, row['job_id']))
            conn.commit()
        children.append(jid)
    return children

def update_worker_status(conn, worker_name, account_name, state, **kw):
    """Persist the latest state, counters, and heartbeat for a worker.

    Parameters
    ----------
    conn : sqlite3.Connection
        Open SQLite connection used for the operation.
    worker_name : str
        Internal worker identifier.
    account_name : str
        Human-readable runtime account label.
    state : str
        Worker state to persist.
    **kw : object
        Additional worker-status fields such as counters, job identifiers, and notes.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    now = utc_now()
    values = {
        'query_id': kw.get('query_id'), 'project_week': kw.get('project_week'),
        'slice_start': kw.get('slice_start'), 'slice_end_exclusive': kw.get('slice_end_exclusive'),
        'current_seen': int(kw.get('current_seen', 0) or 0),
        'current_stored': int(kw.get('current_stored', 0) or 0),
        'session_stored': int(kw.get('session_stored', 0) or 0),
        'session_jobs_done': int(kw.get('session_jobs_done', 0) or 0),
        'session_errors': int(kw.get('session_errors', 0) or 0),
        'started_at_utc': kw.get('started_at_utc') or SESSION_STARTED_UTC or now,
        'note': kw.get('note')
    }
    with DB_LOCK:
        conn.execute("""INSERT INTO worker_heartbeat (
            worker_name, account_name, state, query_id, project_week, slice_start, slice_end_exclusive,
            current_seen, current_stored, session_stored, session_jobs_done, session_errors,
            started_at_utc, last_heartbeat_utc, note
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(worker_name) DO UPDATE SET
            account_name=excluded.account_name, state=excluded.state, query_id=excluded.query_id,
            project_week=excluded.project_week, slice_start=excluded.slice_start,
            slice_end_exclusive=excluded.slice_end_exclusive, current_seen=excluded.current_seen,
            current_stored=excluded.current_stored, session_stored=excluded.session_stored,
            session_jobs_done=excluded.session_jobs_done, session_errors=excluded.session_errors,
            started_at_utc=excluded.started_at_utc, last_heartbeat_utc=excluded.last_heartbeat_utc,
            note=excluded.note""", (
            worker_name, account_name, state, values['query_id'], values['project_week'],
            values['slice_start'], values['slice_end_exclusive'], values['current_seen'],
            values['current_stored'], values['session_stored'], values['session_jobs_done'],
            values['session_errors'], values['started_at_utc'], now, values['note']))
        conn.commit()

def create_db_backup(verbose=False):
    # SQLite online backup while writes are serialized. Rotating one-file backup to limit Drive usage.
    """Create or refresh the rotating SQLite backup file.

    Parameters
    ----------
    verbose : bool
        Whether to print backup status information.

    Returns
    -------
    Path or None
        Backup path when a backup is created, otherwise ``None``.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    tmp = BACKUP_FILE.with_suffix('.tmp.db')
    with DB_LOCK:
        src = connect_db()
        try:
            if tmp.exists():
                tmp.unlink()
            dst = sqlite3.connect(tmp)
            try:
                src.backup(dst)
                dst.commit()
            finally:
                dst.close()
            os.replace(tmp, BACKUP_FILE)
        finally:
            src.close()
    if verbose:
        print(f"Backup updated → {BACKUP_FILE}")
    return BACKUP_FILE

init_database()
seed_initial_jobs()

# -----------------------------------------------------------------------------
# Notebook cell 6
# -----------------------------------------------------------------------------
def iso_z(dt):
    """Convert a datetime value to an ISO 8601 UTC string ending in ``Z``.

    Parameters
    ----------
    dt : datetime
        Timezone-aware datetime value.

    Returns
    -------
    str
        UTC timestamp in ISO 8601 ``Z`` form.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).isoformat(timespec='seconds').replace('+00:00', 'Z')

def project_week_for_timestamp(dt):
    """Map a timestamp to its configured project-week label.

    Parameters
    ----------
    dt : datetime
        Timezone-aware datetime value.

    Returns
    -------
    str
        Project-week label or ``OUT`` when outside the study window.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if dt is None or not (PROJECT_START <= dt < PROJECT_END_EXCLUSIVE):
        return 'OUT', False, False
    idx = (dt.date() - PROJECT_START.date()).days // 7 + 1
    week = f'W{idx:02d}'
    return week, True, (week == 'W21')

def author_hash_from_handle(handle):
    # Web UI often exposes handle but not stable numeric user ID.
    # The resulting pseudonym is stable only while the observed handle is unchanged.
    """Create a salted SHA-256 pseudonym from the available X identity input.

    Parameters
    ----------
    handle : str
        X username or handle.

    Returns
    -------
    str or None
        Salted author hash, or ``None`` when no handle is available.

    Notes
    -----
    X's web UI did not consistently expose a stable numeric author ID during
    historical collection. Most records therefore use a normalized handle;
    records without a handle use the caller's documented content-ID fallback.
    These hashes are for within-X grouping only, not cross-platform linkage.

    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    handle = (handle or '').strip().lstrip('@').lower()
    payload = f"x:handle:{handle}:{AUTHOR_SALT}".encode('utf-8')
    return hashlib.sha256(payload).hexdigest()

def permalink_hash(url):
    """Create a SHA-256 digest for a canonical post URL.

    Parameters
    ----------
    url : str
        URL to inspect or normalize.

    Returns
    -------
    str or None
        SHA-256 URL hash, or ``None`` for an empty URL.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    return hashlib.sha256((url or '').encode('utf-8')).hexdigest() if url else None

def extract_status_id(url):
    """Extract the numeric X post identifier from a status URL.

    Parameters
    ----------
    url : str
        URL to inspect or normalize.

    Returns
    -------
    str or None
        Numeric platform post ID, or ``None`` when no status ID is present.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    m = re.search(r'/status/(\d+)', url or '')
    return m.group(1) if m else None


def _clean_profile_handle_from_url(url):
    """Extract and validate an X profile handle from a profile URL.

    Parameters
    ----------
    url : str
        URL to inspect or normalize.

    Returns
    -------
    str or None
        Validated profile handle, or ``None``.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    try:
        path = urllib.parse.urlparse(url or '').path.strip('/')
        parts = [p for p in path.split('/') if p]
        if not parts:
            return None
        candidate = parts[0]
        blocked = {'home','explore','search','i','settings','notifications','messages','compose'}
        if candidate.lower() in blocked:
            return None
        if re.fullmatch(r'[A-Za-z0-9_]{1,15}', candidate):
            return candidate
    except Exception:
        pass
    return None

def canonical_tweet_url(handle, content_id, observed_url=None):
    """Build the preferred canonical X post URL from available identity fields.

    Parameters
    ----------
    handle : str
        X username or handle.
    content_id : str
        Platform post identifier.
    observed_url : str or None
        Observed post URL, when available.

    Returns
    -------
    str or None
        Canonical post URL, or ``None`` when no post ID exists.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if handle and content_id:
        return f"https://x.com/{handle}/status/{content_id}"
    if content_id:
        return f"https://x.com/i/web/status/{content_id}"
    if observed_url:
        try:
            u = urllib.parse.urlsplit(observed_url)
            return urllib.parse.urlunsplit((u.scheme or 'https', u.netloc or 'x.com', u.path, '', ''))
        except Exception:
            return observed_url
    return None

def extract_author_identity(article, status_link, content_id):
    """Extract the best available author username and display name from a post card.

    Parameters
    ----------
    article : selenium.webdriver.remote.webelement.WebElement
        Rendered X post-card element.
    status_link : str or None
        Observed X status link.
    content_id : str
        Platform post identifier.

    Returns
    -------
    tuple[str | None, str | None]
        Author username and display name.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    handle = None
    display_name = None

    # The status permalink itself usually contains the author handle and is the strongest source.
    try:
        path = urllib.parse.urlparse(status_link or '').path.strip('/')
        parts = [p for p in path.split('/') if p]
        if len(parts) >= 3 and parts[1] == 'status' and parts[2] == str(content_id):
            if re.fullmatch(r'[A-Za-z0-9_]{1,15}', parts[0]):
                handle = parts[0]
    except Exception:
        pass

    try:
        box = article.find_element(By.CSS_SELECTOR, 'div[data-testid="User-Name"]')
        profile_anchor = None
        for a in box.find_elements(By.CSS_SELECTOR, 'a[role="link"]'):
            href = a.get_attribute('href') or ''
            h = _clean_profile_handle_from_url(href)
            if not h:
                continue
            if handle is None:
                handle = h
            if h.lower() == (handle or '').lower():
                profile_anchor = a
                break

        # On current X cards the profile anchor's textContent is normally the display name.
        if profile_anchor is not None:
            txt = re.sub(r'\s+', ' ', profile_anchor.get_attribute('textContent') or '').strip()
            if txt and not txt.startswith('@') and txt.lower() != (handle or '').lower():
                display_name = txt

        # Fallback: inspect visible User-Name lines, excluding handle/date-like text.
        if not display_name:
            lines = [re.sub(r'\s+', ' ', x).strip() for x in (box.text or '').splitlines()]
            for line in lines:
                if not line or line.startswith('@') or line == '·':
                    continue
                if handle and line.lower().lstrip('@') == handle.lower():
                    continue
                if re.fullmatch(r'\d+[smhdwy]', line.lower()):
                    continue
                display_name = line
                break
    except Exception:
        pass

    if handle:
        handle = handle.lstrip('@')
    return handle, display_name

def detect_language_unicode(text):
    """Estimate whether text is Persian, English, or another language using Unicode ranges.

    Parameters
    ----------
    text : str
        Text to inspect.

    Returns
    -------
    tuple[str, float]
        Detected language label and heuristic confidence.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    text = text or ''
    fa = len(re.findall(r'[\u0600-\u06FF]', text))
    en = len(re.findall(r'[A-Za-z]', text))
    total = fa + en
    if total == 0:
        return 'other', 0.0
    if fa > en:
        return 'fa', round(fa / total, 3)
    if en > fa:
        return 'en', round(en / total, 3)
    return 'other', 0.5

def compact_number(value):
    """Parse a compact social metric such as ``1.2K`` into an integer.

    Parameters
    ----------
    value : str or int or float or None
        Compact metric value to parse.

    Returns
    -------
    int or None
        Parsed integer metric, or ``None`` when parsing fails.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if value is None:
        return None
    s = str(value).strip().replace(',', '')
    if not s:
        return None
    mult = 1
    if s[-1:].upper() == 'K':
        mult, s = 1_000, s[:-1]
    elif s[-1:].upper() == 'M':
        mult, s = 1_000_000, s[:-1]
    elif s[-1:].upper() == 'B':
        mult, s = 1_000_000_000, s[:-1]
    try:
        return int(float(s) * mult)
    except Exception:
        nums = re.findall(r'\d[\d,.]*', str(value))
        if not nums:
            return None
        try:
            return int(nums[0].replace(',', ''))
        except Exception:
            return None

def metric_from_element(article, testid):
    """Read and normalize an engagement metric from a tweet-card element.

    Parameters
    ----------
    article : selenium.webdriver.remote.webelement.WebElement
        Rendered X post-card element.
    testid : str
        X ``data-testid`` value for the target metric.

    Returns
    -------
    int or None
        Normalized engagement metric when available.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    try:
        el = article.find_element(By.CSS_SELECTOR, f'[data-testid="{testid}"]')
        for candidate in [(el.get_attribute('aria-label') or ''), (el.text or '')]:
            v = compact_number(candidate)
            if v is not None:
                return v
    except Exception:
        pass
    return None

def infer_content_type(article):
    """Infer the project content type for a rendered X post card.

    Parameters
    ----------
    article : selenium.webdriver.remote.webelement.WebElement
        Rendered X post-card element.

    Returns
    -------
    str
        Project content-type label.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    try:
        social = article.find_elements(By.CSS_SELECTOR, 'div[data-testid="socialContext"]')
        social_text = ' '.join(x.text for x in social if x.text).lower()
        if 'repost' in social_text or 'retweeted' in social_text:
            return 'repost', 'medium'
    except Exception:
        pass
    try:
        if len(article.find_elements(By.CSS_SELECTOR, 'div[data-testid="tweetText"]')) >= 2:
            return 'quote', 'medium'
    except Exception:
        pass
    try:
        if 'replying to' in (article.text or '').lower():
            return 'reply', 'medium'
    except Exception:
        pass
    return 'original_post', 'low'

def create_proxy_extension(host, port, user, pwd, worker_name):
    """Create a temporary Chrome extension for an authenticated HTTP proxy.

    Parameters
    ----------
    host : str
        Proxy host name or IP address supplied at runtime.
    port : str or int
        Proxy port supplied at runtime.
    user : str
        Proxy username supplied at runtime.
    pwd : str
        Proxy password supplied at runtime.
    worker_name : str
        Internal worker identifier.

    Returns
    -------
    Path
        Path to the temporary proxy-extension ZIP file.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if not user or not pwd:
        return None
    ext_dir = Path(f'proxy_ext_{worker_name}')
    ext_dir.mkdir(exist_ok=True)
    manifest = {
        "version": "1.0.0", "manifest_version": 2, "name": f"Proxy-{worker_name}",
        "permissions": ["proxy", "webRequest", "webRequestBlocking", "<all_urls>"],
        "background": {"scripts": ["background.js"]}
    }
    js = (
        'var config={mode:"fixed_servers",rules:{singleProxy:{scheme:"http",host:"%s",port:%s}}};'
        'chrome.proxy.settings.set({value:config,scope:"regular"},function(){});'
        'chrome.webRequest.onAuthRequired.addListener(function(d,cb){cb({authCredentials:{username:"%s",password:"%s"}});},'
        '{urls:["<all_urls>"]},["blocking"]);'
    ) % (host, int(port), user, pwd)
    (ext_dir / 'manifest.json').write_text(json.dumps(manifest), encoding='utf-8')
    (ext_dir / 'background.js').write_text(js, encoding='utf-8')
    import zipfile
    zip_path = Path(f'proxy_{worker_name}.zip')
    with zipfile.ZipFile(zip_path, 'w') as zp:
        zp.write(ext_dir / 'manifest.json', 'manifest.json')
        zp.write(ext_dir / 'background.js', 'background.js')
    return str(zip_path.resolve())

def setup_driver(account, worker_name):
    """Create and configure a Selenium Chrome driver for one worker account.

    Parameters
    ----------
    account : Mapping
        Runtime account configuration loaded from the environment.
    worker_name : str
        Internal worker identifier.

    Returns
    -------
    selenium.webdriver.Chrome
        Configured Chrome driver.

    Raises
    ------
    WebDriverException
        If Chrome cannot be created with the supplied runtime configuration.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    opts = webdriver.ChromeOptions()
    opts.add_argument('--headless=new')
    opts.add_argument('--no-sandbox')
    opts.add_argument('--disable-dev-shm-usage')
    opts.add_argument('--disable-gpu')
    opts.add_argument('--window-size=1440,1200')
    opts.add_argument('--disable-notifications')
    opts.add_argument('--lang=en-US')
    # X is a JavaScript-heavy SPA. Selenium's 'none' strategy prevents driver.get()
    # from blocking on a renderer that is still downloading nonessential resources.
    # safe_get() below performs an explicit DOM usability wait instead.
    opts.page_load_strategy = 'none'

    host, port = account.get('proxy_host'), account.get('proxy_port')
    user, pwd = account.get('proxy_user'), account.get('proxy_pass')
    if host and port:
        if user and pwd:
            ext = create_proxy_extension(host, port, user, pwd, worker_name)
            if ext:
                opts.add_extension(ext)
        else:
            opts.add_argument(f'--proxy-server={host}:{port}')
    if Path('/usr/bin/google-chrome').exists():
        opts.binary_location = '/usr/bin/google-chrome'

    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    driver.set_script_timeout(20)
    return driver


def safe_get(driver, url, ready_timeout=SAFE_GET_READY_TIMEOUT):
    """Navigate to a URL while tolerating slow single-page-application rendering.

    Parameters
    ----------
    driver : selenium.webdriver.Chrome
        Active Selenium Chrome driver.
    url : str
        URL to inspect or normalize.
    ready_timeout : float
        Maximum seconds to wait for a usable page body.

    Returns
    -------
    bool
        ``True`` when a usable page body becomes available.

    Raises
    ------
    WebDriverException
        If the browser session is no longer usable.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    nav_error = None
    try:
        driver.get(url)
    except (TimeoutException, WebDriverException) as e:
        nav_error = e
        logging.warning('Navigation warning for %s: %s', url[:160], str(e).splitlines()[0][:300])
        try:
            driver.execute_script('window.stop();')
        except Exception:
            pass

    deadline = time.monotonic() + max(3, ready_timeout)
    while time.monotonic() < deadline and not STOP_EVENT.is_set():
        try:
            if driver.find_elements(By.TAG_NAME, 'body'):
                return True
        except WebDriverException:
            pass
        time.sleep(0.5)

    # A renderer warning can be harmless if DOM appeared late; do one final check.
    try:
        if driver.find_elements(By.TAG_NAME, 'body'):
            return True
    except Exception:
        pass

    if nav_error:
        logging.warning('safe_get could not obtain a usable DOM after navigation warning.')
    return False


def inject_cookies(driver, cookies):
    # First navigation exists only to establish the x.com cookie domain.
    """Inject the configured X session cookies into a browser session.

    Parameters
    ----------
    driver : selenium.webdriver.Chrome
        Active Selenium Chrome driver.
    cookies : list of dict
        Runtime session-cookie dictionaries.

    Returns
    -------
    bool
        ``True`` when at least one cookie is accepted by Chrome.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if not safe_get(driver, 'https://x.com/', ready_timeout=SAFE_GET_READY_TIMEOUT):
        return False
    time.sleep(1.0)

    added = 0
    for c in cookies or []:
        try:
            cookie = {k: v for k, v in c.items() if k in ['name','value','domain','path','secure','httpOnly']}
            driver.add_cookie(cookie)
            added += 1
        except Exception as e:
            logging.debug('Cookie add failed: %s', e)

    if added == 0:
        logging.warning('No X cookies were accepted by Chrome.')

    # Use an explicit navigation rather than refresh(), which can block on the renderer.
    if not safe_get(driver, 'https://x.com/home', ready_timeout=SAFE_GET_READY_TIMEOUT):
        return False
    time.sleep(3.0)

    try:
        url = (driver.current_url or '').lower()
    except Exception:
        return False
    return not ('/login' in url or '/i/flow/login' in url)

def build_search_url(logical_query, start_date, end_date_exclusive):
    """Build an X live-search URL for one logical query and date slice.

    Parameters
    ----------
    logical_query : str
        Logical query text from the project query registry.
    start_date : datetime or str
        Inclusive search start date.
    end_date_exclusive : datetime or str
        Exclusive search end date.

    Returns
    -------
    tuple[str, str]
        Search URL and exact query string used for provenance.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    query = f"({logical_query}) since:{start_date} until:{end_date_exclusive}"
    return f"https://x.com/search?q={urllib.parse.quote(query)}&src=typed_query&f=live", query

def parse_article(article, query_id, job_week, collection_run_id):
    """Parse one rendered X post card into the project raw-record schema.

    Parameters
    ----------
    article : selenium.webdriver.remote.webelement.WebElement
        Rendered X post-card element.
    query_id : str
        Project query identifier.
    job_week : str
        Project-week label associated with the parsed post.
    collection_run_id : str
        Identifier of the current collection run.

    Returns
    -------
    dict or None
        Parsed raw record, or ``None`` when the card lacks a usable post ID.

    Raises
    ------
    WebDriverException
        If Selenium cannot inspect the rendered post card.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    link = None
    for a in article.find_elements(By.CSS_SELECTOR, 'a[href*="/status/"]'):
        href = a.get_attribute('href')
        if extract_status_id(href):
            link = href
            break
    content_id = extract_status_id(link)
    if not content_id:
        return None

    text_raw = ''
    language_reported = None
    try:
        text_el = article.find_element(By.CSS_SELECTOR, 'div[data-testid="tweetText"]')
        text_raw = text_el.get_attribute('textContent')
        language_reported = text_el.get_attribute('lang')
    except Exception:
        pass

    handle, display_name = extract_author_identity(article, link, content_id)
    tweet_url = canonical_tweet_url(handle, content_id, link)

    if not handle:
        hash_input = f'unknown:{content_id}'
        hash_method = 'content_id_fallback_v1'
    else:
        hash_input = handle
        hash_method = 'handle_fallback_v1'

    created_dt = None
    try:
        raw = article.find_element(By.CSS_SELECTOR, 'time[datetime]').get_attribute('datetime')
        created_dt = datetime.fromisoformat(raw.replace('Z', '+00:00')).astimezone(timezone.utc)
    except Exception:
        pass

    actual_week, in_window, is_partial = project_week_for_timestamp(created_dt)
    record_week = actual_week if created_dt else job_week
    if created_dt is None:
        in_window = job_week != 'OUT'
        is_partial = job_week == 'W21'

    content_type, content_type_conf = infer_content_type(article)
    lang_detected, lang_conf = detect_language_unicode(text_raw)
    now = utc_now()

    return {
        'platform': 'x',
        'platform_content_id': content_id,
        'content_type': content_type,
        'created_at_utc': iso_z(created_dt),
        'collected_at_utc': now,
        'text_raw': text_raw,
        'author_username': handle,
        'author_display_name': display_name,
        'tweet_url': tweet_url,
        'author_hash': author_hash_from_handle(hash_input),
        'project_week': record_week,
        'job_project_week': job_week,
        'in_window': int(bool(in_window)),
        'is_partial_week': int(bool(is_partial)),
        'query_id': query_id,
        'collection_run_id': collection_run_id,
        'source_id': None,
        'source_container': 'x_search',
        'source_container_id': None,
        'source_parent_id': None,
        'source_parent_title': None,
        'parent_id': None,
        'query_version': PROJECT_CONFIG['query_version'],
        'discovery_route': QUERY_BY_ID[query_id]['route'],
        'matched_query_ids': query_id,
        'collector_version': PROJECT_CONFIG['collector_version'],
        'permalink_hash': permalink_hash(link),
        'source_total_available': None,
        'sampling_applied': 0,
        'items_kept': None,
        'random_seed': None,
        'engagement_score': metric_from_element(article, 'like'),
        'engagement_replies': metric_from_element(article, 'reply'),
        'engagement_shares': metric_from_element(article, 'retweet'),
        'engagement_quotes': None,
        'engagement_views': None,
        'engagement_collected_at_utc': now,
        'author_is_verified': None,
        'author_follower_count': None,
        'author_account_age_days': None,
        'automation_risk_score': None,
        'language_reported': language_reported,
        'language_detected': lang_detected,
        'language_confidence': lang_conf,
        'content_status': 'active',
        'geo_method': None,
        'country_or_region': None,
        'geo_confidence': None,
        'geo_granularity': None,
        'geo_limitations': None,
        'author_hash_method': hash_method,
        'content_type_confidence': content_type_conf,
    }

RAW_COLUMNS = [
    'platform','platform_content_id','content_type','created_at_utc','collected_at_utc','text_raw',
    'author_username','author_display_name','tweet_url','author_hash',
    'project_week','in_window','is_partial_week','query_id','collection_run_id','source_id','source_container',
    'source_container_id','source_parent_id','source_parent_title','parent_id','query_version','discovery_route',
    'matched_query_ids','collector_version','permalink_hash','source_total_available','sampling_applied','items_kept',
    'random_seed','engagement_score','engagement_replies','engagement_shares','engagement_quotes','engagement_views',
    'engagement_collected_at_utc','author_is_verified','author_follower_count','author_account_age_days',
    'automation_risk_score','language_reported','language_detected','language_confidence','content_status','geo_method',
    'country_or_region','geo_confidence','geo_granularity','geo_limitations','author_hash_method','content_type_confidence'
]


def merge_query_ids(existing, new_id):
    """Merge a query identifier into a semicolon-delimited unique query-id list.

    Parameters
    ----------
    existing : str or None
        Existing semicolon-delimited query-id string.
    new_id : str
        Query identifier to add.

    Returns
    -------
    str
        Semicolon-delimited unique query identifiers.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    ids = [x for x in (existing or '').split(';') if x]
    if new_id not in ids:
        ids.append(new_id)
    return ';'.join(ids)

def _journal_lock(name):
    """Return the per-worker lock used to serialize recovery-journal writes.

    Parameters
    ----------
    name : str
        Worker name used to address an in-process lock.

    Returns
    -------
    threading.RLock
        Per-worker recovery-journal lock.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    with STATE_LOCK:
        if name not in JOURNAL_LOCKS:
            JOURNAL_LOCKS[name] = threading.Lock()
        return JOURNAL_LOCKS[name]

def recovery_journal_path(worker_name):
    """Return the recovery-journal file path for a worker.

    Parameters
    ----------
    worker_name : str
        Internal worker identifier.

    Returns
    -------
    Path
        Recovery-journal path for the worker.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    safe = re.sub(r'[^A-Za-z0-9_-]+', '-', worker_name)
    return RECOVERY_DIR / f'{safe}_tweets.jsonl'

def append_recovery_record(worker_name, rec):
    """Append one recoverable record to a worker journal before database commit.

    Parameters
    ----------
    worker_name : str
        Internal worker identifier.
    rec : Mapping
        Parsed raw record to journal or persist.

    Returns
    -------
    tuple[str, int]
        Journal filename and end-byte offset after the appended line.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    path = recovery_journal_path(worker_name)
    payload = {
        'kind': 'tweet_record_v2',
        'journaled_at_utc': utc_now(),
        'record': {k: rec.get(k) for k in (RAW_COLUMNS + ['job_project_week'])}
    }
    line = (json.dumps(payload, ensure_ascii=False, separators=(',', ':')) + '\n').encode('utf-8')
    lock = _journal_lock(path.name)
    with lock:
        with open(path, 'ab', buffering=0) as f:
            f.write(line)
            f.flush()
            if RECOVERY_FSYNC_EVERY == 1:
                os.fsync(f.fileno())
            end_offset = f.tell()
    return path.name, end_offset

def store_record(conn, rec, journal_file=None, journal_offset=None):
    """Insert or merge one raw record and acknowledge its journal offset.

    Parameters
    ----------
    conn : sqlite3.Connection
        Open SQLite connection used for the operation.
    rec : Mapping
        Parsed raw record to journal or persist.
    journal_file : str or Path
        Recovery-journal file containing the record.
    journal_offset : int
        End-byte offset to acknowledge after commit.

    Returns
    -------
    bool
        ``True`` when a new raw row is inserted; ``False`` when an existing row is merged.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    placeholders = ','.join('?' for _ in RAW_COLUMNS)
    cols = ','.join(RAW_COLUMNS)
    values = [rec.get(c) for c in RAW_COLUMNS]
    inserted = False
    with DB_LOCK:
        try:
            conn.execute(f"INSERT INTO tweets_raw ({cols}) VALUES ({placeholders})", values)
            inserted = True
        except sqlite3.IntegrityError:
            row = conn.execute(
                "SELECT matched_query_ids FROM tweets_raw WHERE platform=? AND platform_content_id=?",
                (rec['platform'], rec['platform_content_id'])
            ).fetchone()
            merged = merge_query_ids(row['matched_query_ids'] if row else '', rec['query_id'])
            conn.execute("""UPDATE tweets_raw
                            SET matched_query_ids=?,
                                author_username=COALESCE(NULLIF(?, ''), author_username),
                                author_display_name=COALESCE(NULLIF(?, ''), author_display_name),
                                tweet_url=COALESCE(NULLIF(?, ''), tweet_url),
                                engagement_score=COALESCE(?, engagement_score),
                                engagement_replies=COALESCE(?, engagement_replies),
                                engagement_shares=COALESCE(?, engagement_shares),
                                engagement_collected_at_utc=?
                            WHERE platform=? AND platform_content_id=?""",
                         (merged, rec.get('author_username'), rec.get('author_display_name'), rec.get('tweet_url'),
                          rec['engagement_score'], rec['engagement_replies'], rec['engagement_shares'],
                          rec['engagement_collected_at_utc'], rec['platform'], rec['platform_content_id']))

        conn.execute("""INSERT OR IGNORE INTO tweet_matches
            (platform, platform_content_id, query_id, project_week, collection_run_id, matched_at_utc)
            VALUES (?, ?, ?, ?, ?, ?)""",
            (rec['platform'], rec['platform_content_id'], rec['query_id'], rec.get('job_project_week', rec['project_week']),
             rec['collection_run_id'], rec['collected_at_utc']))

        if journal_file is not None and journal_offset is not None:
            conn.execute("""INSERT INTO recovery_offsets(journal_file, byte_offset, updated_at_utc)
                          VALUES(?,?,?)
                          ON CONFLICT(journal_file) DO UPDATE SET
                          byte_offset=excluded.byte_offset, updated_at_utc=excluded.updated_at_utc""",
                         (journal_file, int(journal_offset), utc_now()))
        conn.commit()
    return inserted

def persist_record(conn, rec, worker_name):
    # Journal-first, DB-second = best-effort crash durability on Colab/Drive.
    """Journal and persist one raw record using the durability protocol.

    Parameters
    ----------
    conn : sqlite3.Connection
        Open SQLite connection used for the operation.
    rec : Mapping
        Parsed raw record to journal or persist.
    worker_name : str
        Internal worker identifier.

    Returns
    -------
    bool
        ``True`` when a new raw row is inserted; otherwise ``False``.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    jf, off = append_recovery_record(worker_name, rec)
    return store_record(conn, rec, journal_file=jf, journal_offset=off)

def recover_from_journals():
    """Replay unacknowledged recovery-journal records into SQLite idempotently.

    Returns
    -------
    int
        Number of journal records replayed during recovery.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    replayed = 0
    bad_lines = 0
    for path in sorted(RECOVERY_DIR.glob('*_tweets.jsonl')):
        row = conn.execute('SELECT byte_offset FROM recovery_offsets WHERE journal_file=?', (path.name,)).fetchone()
        start_offset = int(row['byte_offset']) if row else 0
        size = path.stat().st_size
        if start_offset > size:
            # Journal was replaced/truncated; conservatively replay from beginning.
            start_offset = 0
        if start_offset == size:
            continue
        with open(path, 'rb') as f:
            f.seek(start_offset)
            while True:
                pos_before = f.tell()
                line = f.readline()
                if not line:
                    break
                end_pos = f.tell()
                try:
                    payload = json.loads(line.decode('utf-8'))
                    if payload.get('kind') not in {'tweet_record_v1', 'tweet_record_v2'}:
                        continue
                    rec = payload['record']
                    store_record(conn, rec, journal_file=path.name, journal_offset=end_pos)
                    replayed += 1
                except Exception as e:
                    bad_lines += 1
                    logging.warning('Recovery stopped at %s byte %s: %s', path.name, pos_before, e)
                    # Partial final line may be from a hard kill; do not advance offset past it.
                    break
    conn.close()
    if replayed or bad_lines:
        print(f"Recovery replay: {replayed:,} record(s); unreadable/partial line(s): {bad_lines}")
    else:
        print('Recovery journals are fully synchronized with SQLite.')
    return replayed

# -----------------------------------------------------------------------------
# Notebook cell 7
# -----------------------------------------------------------------------------
class GracefulStop(Exception):
    pass

class SessionInvalid(Exception):
    pass

class TemporaryAccessIssue(Exception):
    pass

def make_collection_run_id(job_row, worker_name):
    """Create a unique identifier for one worker execution of a collection job.

    Parameters
    ----------
    job_row : sqlite3.Row or Mapping
        Persisted collection job row.
    worker_name : str
        Internal worker identifier.

    Returns
    -------
    str
        Unique collection-run identifier.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S%fZ')
    safe_worker = re.sub(r'[^A-Za-z0-9_-]+', '-', worker_name)
    return f"XRUN-{job_row['query_id']}-{job_row['project_week']}-D{job_row['depth']}-{safe_worker}-{stamp}"

def should_split_job(job_row, returned_unique, last_scroll_new):
    """Decide whether a date slice is dense enough to require adaptive splitting.

    Parameters
    ----------
    job_row : sqlite3.Row or Mapping
        Persisted collection job row.
    returned_unique : int
        Number of unique posts observed in the current slice.
    last_scroll_new : int
        Number of newly observed posts from the last scroll.

    Returns
    -------
    bool
        Whether the current job should be replaced by child date slices.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    start = datetime.fromisoformat(job_row['slice_start']).replace(tzinfo=timezone.utc)
    end = datetime.fromisoformat(job_row['slice_end_exclusive']).replace(tzinfo=timezone.utc)
    days = (end - start).days
    if days <= MIN_SLICE_DAYS:
        return False
    return last_scroll_new >= SPLIT_LAST_SCROLL_NEW or returned_unique >= SPLIT_TOTAL_SEEN

def detect_access_state(driver):
    """Classify visible browser state without bypassing access or verification controls.

    Parameters
    ----------
    driver : selenium.webdriver.Chrome
        Active Selenium Chrome driver.

    Returns
    -------
    tuple[str, str]
        Detected access state and explanatory note.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    url = (driver.current_url or '').lower()
    if '/login' in url or '/i/flow/login' in url:
        return 'SESSION_INVALID', 'X redirected this session to login.'
    try:
        body = (driver.find_element(By.TAG_NAME, 'body').text or '').lower()[:25000]
    except Exception:
        body = ''
    verification_terms = ['captcha', 'verify your identity', 'unusual activity', 'confirm your identity']
    if any(t in body for t in verification_terms):
        return 'VERIFICATION_REQUIRED', 'Manual account verification/CAPTCHA appears to be required.'
    limit_terms = ['rate limit', 'try again later', 'something went wrong. try reloading']
    if any(t in body for t in limit_terms):
        return 'TEMP_LIMIT', 'X returned a temporary limitation/error page.'
    return 'OK', ''

def scrape_job(driver, conn, job_row, worker_name, account_name, stats):
    """Execute one query/date-slice collection job and persist its audit metadata.

    Parameters
    ----------
    driver : selenium.webdriver.Chrome
        Active Selenium Chrome driver.
    conn : sqlite3.Connection
        Open SQLite connection used for the operation.
    job_row : sqlite3.Row or Mapping
        Persisted collection job row.
    worker_name : str
        Internal worker identifier.
    account_name : str
        Human-readable runtime account label.
    stats : dict
        Mutable per-worker counters for the current session.

    Returns
    -------
    dict
        Per-job collection metrics and audit information.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if STOP_EVENT.is_set():
        raise GracefulStop()

    q = QUERY_BY_ID.get(job_row['query_id'])
    if not q:
       logging.warning('Job %s has unknown query_id %s; skipping.', job_row['job_id'], job_row['query_id'])
       set_job_status(conn, job_row['job_id'], 'failed', 'query_not_in_active_registry')
       return
    run_id = make_collection_run_id(job_row, worker_name)
    started = utc_now()
    url, exact_query = build_search_url(q['logical_query'], job_row['slice_start'], job_row['slice_end_exclusive'])

    error_count = 0
    stored_count = 0
    unique_seen = set()
    oldest = None
    newest = None
    last_scroll_new = 0
    no_new_streak = 0
    scrolls_used = 0

    update_worker_status(conn, worker_name, account_name, 'LOADING',
        query_id=q['query_id'], project_week=job_row['project_week'],
        slice_start=job_row['slice_start'], slice_end_exclusive=job_row['slice_end_exclusive'],
        session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
        started_at_utc=stats['started_at'], note='Opening X search')

    start_dt = datetime.fromisoformat(job_row['slice_start']).replace(tzinfo=timezone.utc)
    end_dt = datetime.fromisoformat(job_row['slice_end_exclusive']).replace(tzinfo=timezone.utc)
    days = (end_dt - start_dt).days
    max_scrolls = MAX_SCROLLS_MIN_DAY if days <= MIN_SLICE_DAYS else MAX_SCROLLS_PER_SLICE

    if not safe_get(driver, url, ready_timeout=SAFE_GET_READY_TIMEOUT):
        error_count += 1
        stats['errors'] += 1
        raise TemporaryAccessIssue('Search navigation did not produce a usable DOM')

    if STOP_EVENT.is_set():
        raise GracefulStop()

    time.sleep(random.uniform(2.5, 4.5))
    access_state, access_note = detect_access_state(driver)
    if access_state in {'SESSION_INVALID', 'VERIFICATION_REQUIRED'}:
        update_worker_status(conn, worker_name, account_name, access_state,
            query_id=q['query_id'], project_week=job_row['project_week'],
            slice_start=job_row['slice_start'], slice_end_exclusive=job_row['slice_end_exclusive'],
            session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
            started_at_utc=stats['started_at'], note=access_note)
        raise SessionInvalid(access_note)
    if access_state == 'TEMP_LIMIT':
        update_worker_status(conn, worker_name, account_name, 'BACKOFF',
            query_id=q['query_id'], project_week=job_row['project_week'],
            slice_start=job_row['slice_start'], slice_end_exclusive=job_row['slice_end_exclusive'],
            session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
            started_at_utc=stats['started_at'], note=access_note)
        raise TemporaryAccessIssue(access_note)

    for scroll_idx in range(max_scrolls):
        if STOP_EVENT.is_set():
            raise GracefulStop()
        scrolls_used = scroll_idx + 1
        before = len(unique_seen)
        try:
            articles = driver.find_elements(By.CSS_SELECTOR, 'article[data-testid="tweet"]')
        except Exception:
            articles = []
            error_count += 1
            stats['errors'] += 1

        for article in articles:
            if STOP_EVENT.is_set():
                raise GracefulStop()
            try:
                rec = parse_article(article, q['query_id'], job_row['project_week'], run_id)
                if not rec:
                    continue
                pid = rec['platform_content_id']
                if pid in unique_seen:
                    continue
                unique_seen.add(pid)
                if persist_record(conn, rec, worker_name):
                    stored_count += 1
                    stats['stored'] += 1
                if rec['created_at_utc']:
                    oldest = min(oldest, rec['created_at_utc']) if oldest else rec['created_at_utc']
                    newest = max(newest, rec['created_at_utc']) if newest else rec['created_at_utc']
            except GracefulStop:
                raise
            except Exception as e:
                error_count += 1
                stats['errors'] += 1
                logging.debug('%s parse/store error: %s', worker_name, e)

        last_scroll_new = len(unique_seen) - before
        no_new_streak = no_new_streak + 1 if last_scroll_new == 0 else 0
        update_worker_status(conn, worker_name, account_name, 'SCRAPING',
            query_id=q['query_id'], project_week=job_row['project_week'],
            slice_start=job_row['slice_start'], slice_end_exclusive=job_row['slice_end_exclusive'],
            current_seen=len(unique_seen), current_stored=stored_count,
            session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
            started_at_utc=stats['started_at'], note=f'Scroll {scrolls_used}/{max_scrolls} · last new={last_scroll_new}')

        if no_new_streak >= NO_NEW_SCROLL_LIMIT:
            break
        try:
            driver.execute_script("window.scrollBy(0, Math.max(window.innerHeight * 0.9, 800));")
        except Exception:
            error_count += 1
            stats['errors'] += 1
        time.sleep(random.uniform(SCROLL_MIN_DELAY, SCROLL_MAX_DELAY))

    if STOP_EVENT.is_set():
        raise GracefulStop()

    split = should_split_job(job_row, len(unique_seen), last_scroll_new)
    children = enqueue_split_children(conn, job_row) if split else []

    records_in_window = conn.execute("""SELECT COUNT(DISTINCT m.platform_content_id)
        FROM tweet_matches m JOIN tweets_raw t
        ON t.platform=m.platform AND t.platform_content_id=m.platform_content_id
        WHERE m.collection_run_id=? AND t.in_window=1""", (run_id,)).fetchone()[0]

    finished = utc_now()
    notes = json.dumps({
        'exact_search_query': exact_query,
        'children_created': children,
        'dense_day_scroll_limit_used': days <= MIN_SLICE_DAYS,
    }, ensure_ascii=False)

    with DB_LOCK:
        conn.execute("""INSERT OR REPLACE INTO x_subruns (
            collection_run_id, job_id, worker_name, account_name, platform, query_id, query_text, query_version,
            project_week, discovery_route, source_id, sort_mode, slice_start, slice_end_exclusive, depth,
            started_at_utc, finished_at_utc, returned_count, stored_count, records_in_window,
            oldest_record_utc, newest_record_utc, error_count, last_scroll_new, scrolls_used, split_triggered, notes
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
            run_id, job_row['job_id'], worker_name, account_name, 'x', q['query_id'], q['logical_query'],
            PROJECT_CONFIG['query_version'], job_row['project_week'], q['route'], None, PROJECT_CONFIG['sort_mode'],
            job_row['slice_start'], job_row['slice_end_exclusive'], job_row['depth'], started, finished,
            len(unique_seen), stored_count, records_in_window, oldest, newest, error_count,
            last_scroll_new, scrolls_used, int(bool(children)), notes
        ))
        conn.commit()

    set_job_status(conn, job_row['job_id'], 'split' if children else 'done')
    stats['jobs'] += 1
    update_worker_status(conn, worker_name, account_name, 'IDLE',
        session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
        started_at_utc=stats['started_at'], note='Job committed to Drive')
    logging.info('%s %s %s %s→%s seen=%s new=%s split=%s',
                 worker_name, q['query_id'], job_row['project_week'], job_row['slice_start'],
                 job_row['slice_end_exclusive'], len(unique_seen), stored_count, bool(children))

def worker_loop(worker_index, account):
    """Run the lifecycle of one independent browser worker until completion or stop.

    Parameters
    ----------
    worker_index : int
        Zero-based worker index.
    account : Mapping
        Runtime account configuration loaded from the environment.

    Returns
    -------
    str
        Final worker state.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    worker_name = f'worker-{worker_index+1}'
    account_name = account.get('name', worker_name)
    conn = connect_db()
    driver = None
    stats = {'stored': 0, 'jobs': 0, 'errors': 0, 'started_at': utc_now()}
    current_job = None
    terminal_state = None
    terminal_note = None

    try:
        # Avoid launching three Chrome renderers through three proxies in the same instant.
        stagger = worker_index * STARTUP_STAGGER_SECONDS
        if stagger:
            update_worker_status(conn, worker_name, account_name, 'STARTUP_WAIT',
                                 started_at_utc=stats['started_at'],
                                 note=f'Staggering Chrome startup by {stagger}s')
            for _ in range(stagger):
                if STOP_EVENT.is_set():
                    raise GracefulStop()
                time.sleep(1)

        startup_ok = False
        startup_error = None
        for attempt in range(1, STARTUP_MAX_ATTEMPTS + 1):
            if STOP_EVENT.is_set():
                raise GracefulStop()
            try:
                update_worker_status(conn, worker_name, account_name,
                                     'STARTING' if attempt == 1 else 'STARTUP_RETRY',
                                     started_at_utc=stats['started_at'],
                                     session_errors=stats['errors'],
                                     note=f'Chrome/session startup attempt {attempt}/{STARTUP_MAX_ATTEMPTS}')
                if driver is not None:
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    driver = None
                driver = setup_driver(account, worker_name)
                if not inject_cookies(driver, account.get('cookies')):
                    raise SessionInvalid(f'Login/session validation failed for {account_name}')
                startup_ok = True
                break
            except GracefulStop:
                raise
            except Exception as e:
                startup_error = e
                stats['errors'] += 1
                logging.warning('%s startup attempt %s/%s failed: %s',
                                worker_name, attempt, STARTUP_MAX_ATTEMPTS,
                                str(e).splitlines()[0][:500])
                update_worker_status(conn, worker_name, account_name, 'STARTUP_RETRY',
                                     session_stored=stats['stored'], session_jobs_done=stats['jobs'],
                                     session_errors=stats['errors'], started_at_utc=stats['started_at'],
                                     note=f'Attempt {attempt} failed: {str(e).splitlines()[0][:220]}')
                try:
                    if driver is not None:
                        driver.quit()
                except Exception:
                    pass
                driver = None
                if attempt < STARTUP_MAX_ATTEMPTS:
                    for _ in range(STARTUP_RETRY_SECONDS):
                        if STOP_EVENT.is_set():
                            raise GracefulStop()
                        time.sleep(1)

        if not startup_ok:
            terminal_state = 'STARTUP_FAILED'
            terminal_note = f'Session could not start after {STARTUP_MAX_ATTEMPTS} attempts: {str(startup_error)[:260]}'
            update_worker_status(conn, worker_name, account_name, terminal_state,
                                 session_stored=stats['stored'], session_jobs_done=stats['jobs'],
                                 session_errors=stats['errors'], started_at_utc=stats['started_at'],
                                 note=terminal_note)
            logging.error('%s startup failed; no job was claimed.', worker_name)
            return

        update_worker_status(conn, worker_name, account_name, 'READY',
                             session_stored=stats['stored'], session_jobs_done=stats['jobs'],
                             session_errors=stats['errors'], started_at_utc=stats['started_at'],
                             note='Session validated')
        logging.info('%s started with %s', worker_name, account_name)

        while not STOP_EVENT.is_set():
            row = claim_next_job(conn, worker_name)
            current_job = row
            if row is None:
                terminal_state = 'DONE'
                terminal_note = 'No open jobs remain for this worker'
                break
            try:
                scrape_job(driver, conn, row, worker_name, account_name, stats)
                current_job = None
            except GracefulStop:
                set_job_status(conn, row['job_id'], 'pending', 'graceful_stop')
                terminal_state = 'STOPPED'
                terminal_note = 'Graceful stop requested; in-flight job returned to pending'
                update_worker_status(conn, worker_name, account_name, 'STOPPING',
                    session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
                    started_at_utc=stats['started_at'], note=terminal_note)
                break
            except SessionInvalid as e:
                stats['errors'] += 1
                set_job_status(conn, row['job_id'], 'pending', str(e)[:1000])
                terminal_state = 'VERIFICATION_REQUIRED'
                terminal_note = str(e)[:300]
                update_worker_status(conn, worker_name, account_name, terminal_state,
                    query_id=row['query_id'], project_week=row['project_week'],
                    slice_start=row['slice_start'], slice_end_exclusive=row['slice_end_exclusive'],
                    session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
                    started_at_utc=stats['started_at'], note=terminal_note)
                logging.error('%s stopped for manual session/account verification: %s', worker_name, e)
                break
            except TemporaryAccessIssue as e:
                stats['errors'] += 1
                status = 'pending' if row['attempt_count'] < MAX_JOB_ATTEMPTS else 'failed'
                set_job_status(conn, row['job_id'], status, str(e)[:1000])
                update_worker_status(conn, worker_name, account_name, 'BACKOFF',
                    query_id=row['query_id'], project_week=row['project_week'],
                    slice_start=row['slice_start'], slice_end_exclusive=row['slice_end_exclusive'],
                    session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
                    started_at_utc=stats['started_at'], note=f'Temporary navigation/X issue; sleeping {JOB_BACKOFF_SECONDS}s')
                for _ in range(JOB_BACKOFF_SECONDS):
                    if STOP_EVENT.is_set():
                        break
                    time.sleep(1)
                current_job = None
            except Exception as e:
                stats['errors'] += 1
                logging.exception('%s job failed: %s', worker_name, row['job_id'])
                status = 'pending' if row['attempt_count'] < MAX_JOB_ATTEMPTS else 'failed'
                set_job_status(conn, row['job_id'], status, str(e)[:1000])
                update_worker_status(conn, worker_name, account_name, 'ERROR/BACKOFF',
                    query_id=row['query_id'], project_week=row['project_week'],
                    slice_start=row['slice_start'], slice_end_exclusive=row['slice_end_exclusive'],
                    session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
                    started_at_utc=stats['started_at'], note=str(e)[:300])
                for _ in range(JOB_BACKOFF_SECONDS):
                    if STOP_EVENT.is_set():
                        break
                    time.sleep(1)
                current_job = None

        if terminal_state is None:
            terminal_state = 'STOPPED' if STOP_EVENT.is_set() else 'DONE'
            terminal_note = 'Worker loop ended'

    except GracefulStop:
        terminal_state = 'STOPPED'
        terminal_note = 'Stopped during startup before claiming a job'
    except Exception as e:
        stats['errors'] += 1
        terminal_state = 'WORKER_FAILED'
        terminal_note = str(e)[:300]
        logging.exception('%s fatal worker error', worker_name)
    finally:
        if current_job is not None:
            try:
                set_job_status(conn, current_job['job_id'], 'pending', 'worker_shutdown')
            except Exception:
                pass
        try:
            if driver:
                driver.quit()
        except Exception:
            pass
        if terminal_state is None:
            terminal_state = 'STOPPED' if STOP_EVENT.is_set() else 'WORKER_FAILED'
            terminal_note = terminal_note or 'Worker ended without terminal state'
        try:
            update_worker_status(conn, worker_name, account_name, terminal_state,
                session_stored=stats['stored'], session_jobs_done=stats['jobs'], session_errors=stats['errors'],
                started_at_utc=stats['started_at'], note=terminal_note or terminal_state)
        except Exception:
            pass
        conn.close()
        logging.info('%s ended with state=%s', worker_name, terminal_state)

def run_collection(max_workers=MAX_WORKERS):
    """Run the multi-worker collection controller with graceful stop and recovery.

    Parameters
    ----------
    max_workers : int
        Maximum number of concurrent browser workers.

    Returns
    -------
    None
        This controller reports progress and persists results as side effects.

    Raises
    ------
    RuntimeError
        If no runtime account configuration is available.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    global SESSION_STARTED_MONO, SESSION_STARTED_UTC, LAST_BACKUP_MONO
    if not ACCOUNTS:
        raise RuntimeError('ACCOUNTS is empty.')

    STOP_EVENT.clear()
    reset_stale_running_jobs()
    recover_from_journals()
    seed_initial_jobs(reset_stale_running=False)

    workers = min(max_workers, len(ACCOUNTS), 3)
    SESSION_STARTED_MONO = time.monotonic()
    SESSION_STARTED_UTC = utc_now()
    LAST_BACKUP_MONO = SESSION_STARTED_MONO
    session_id = 'XSESSION-' + datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')

    conn = connect_db()
    start_raw = conn.execute('SELECT COUNT(*) FROM tweets_raw').fetchone()[0]
    with DB_LOCK:
        conn.execute("""INSERT OR REPLACE INTO collector_sessions
            (session_id, started_at_utc, status, workers, start_raw_count)
            VALUES(?,?,?,?,?)""", (session_id, SESSION_STARTED_UTC, 'running', workers, start_raw))
        conn.execute('DELETE FROM worker_heartbeat')
        conn.commit()
    conn.close()

    print(f"Starting {workers} independent workers — one browser/session per account (stagger={STARTUP_STAGGER_SECONDS}s).")
    print("Use Colab 'Interrupt execution' / Ctrl+C for graceful stop. Re-run this cell later to resume.")
    print("Adaptive slicing minimum = 1 day; dense single-day slices use a slightly deeper scroll cap.")

    ex = ThreadPoolExecutor(max_workers=workers, thread_name_prefix='xworker')
    futures = [ex.submit(worker_loop, i, ACCOUNTS[i]) for i in range(workers)]
    handle = display(HTML(progress_dashboard_html()), display_id=True)
    interrupted = False

    try:
        while not all(f.done() for f in futures):
            try:
                handle.update(HTML(progress_dashboard_html()))
            except Exception:
                pass
            if time.monotonic() - LAST_BACKUP_MONO >= BACKUP_INTERVAL_SECONDS:
                try:
                    create_db_backup(verbose=False)
                    LAST_BACKUP_MONO = time.monotonic()
                except Exception as e:
                    logging.warning('Periodic backup failed: %s', e)
            time.sleep(DASHBOARD_REFRESH_SECONDS)
    except KeyboardInterrupt:
        interrupted = True
        STOP_EVENT.set()
        print("\nStop requested. Workers are returning in-flight jobs to pending and closing cleanly...")
    finally:
        if interrupted:
            STOP_EVENT.set()
        # Workers check STOP_EVENT between records/scrolls; page-load timeout is bounded.
        for f in futures:
            try:
                f.result()
            except Exception as e:
                logging.error('Worker future ended with error: %s', e)
        ex.shutdown(wait=True, cancel_futures=False)
        reset_stale_running_jobs()

        try:
            recover_from_journals()
        except Exception as e:
            logging.warning('Final recovery sync failed: %s', e)
        try:
            create_db_backup(verbose=True)
        except Exception as e:
            logging.warning('Final DB backup failed: %s', e)

        conn = connect_db()
        end_raw = conn.execute('SELECT COUNT(*) FROM tweets_raw').fetchone()[0]
        session_status = 'stopped_by_user' if interrupted else 'completed'
        with DB_LOCK:
            conn.execute("""UPDATE collector_sessions SET finished_at_utc=?, status=?, end_raw_count=?
                          WHERE session_id=?""", (utc_now(), session_status, end_raw, session_id))
            conn.commit()
        conn.close()

        try:
            handle.update(HTML(progress_dashboard_html(final=True)))
        except Exception:
            pass

        # Export only at controlled shutdown/completion; SQLite + journals are the live source of truth.
        try:
            export_all()
        except Exception as e:
            logging.warning('Final export failed; DB/journals are still preserved: %s', e)

    if interrupted:
        print('Graceful stop complete. Re-run run_collection(...) later to resume from Drive.')
    else:
        print('Collection workers finished.')

# -----------------------------------------------------------------------------
# Notebook cell 8
# -----------------------------------------------------------------------------
def job_stats():
    """Print and return a compact summary of persisted queue and raw-record counts.

    Returns
    -------
    pandas.DataFrame
        Job-status summary data frame.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    df = pd.read_sql_query("SELECT status, COUNT(*) AS jobs FROM x_jobs GROUP BY status ORDER BY status", conn)
    raw = conn.execute('SELECT COUNT(*) FROM tweets_raw').fetchone()[0]
    matches = conn.execute('SELECT COUNT(*) FROM tweet_matches').fetchone()[0]
    conn.close()
    print(df.to_string(index=False))
    print(f"Raw unique tweets: {raw:,}")
    print(f"Query/run matches:  {matches:,}")
    return df

def _fmt_duration(seconds):
    """Format an elapsed duration in a compact human-readable form.

    Parameters
    ----------
    seconds : float or int or None
        Duration in seconds.

    Returns
    -------
    str
        Compact formatted duration.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if seconds is None or seconds < 0:
        return '—'
    seconds = int(seconds)
    d, rem = divmod(seconds, 86400)
    h, rem = divmod(rem, 3600)
    m, s = divmod(rem, 60)
    if d:
        return f'{d}d {h:02d}h {m:02d}m'
    if h:
        return f'{h}h {m:02d}m'
    return f'{m}m {s:02d}s'

def _pct(n, d):
    """Calculate a percentage constrained to the inclusive range 0–100.

    Parameters
    ----------
    n : float or int
        Numerator value.
    d : float or int
        Denominator value.

    Returns
    -------
    float
        Percentage in the range 0–100.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    return 0.0 if not d else max(0.0, min(100.0, 100.0 * n / d))

def _badge(state):
    """Render an HTML status badge for a worker state.

    Parameters
    ----------
    state : str
        Worker state to persist.

    Returns
    -------
    str
        HTML fragment containing the status badge.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    state = state or 'UNKNOWN'
    cls = 'gray'
    if state in {'SCRAPING','READY','IDLE','LOADING'}: cls='green'
    elif state in {'BACKOFF','ERROR/BACKOFF','TEMP_LIMIT'}: cls='amber'
    elif state in {'VERIFICATION_REQUIRED','SESSION_INVALID'}: cls='red'
    elif state in {'DONE','STOPPED'}: cls='blue'
    return f'<span class="badge {cls}">{html.escape(state)}</span>'

def progress_dashboard_html(final=False):
    """Build the live HTML dashboard from persisted collection state.

    Parameters
    ----------
    final : bool
        Whether the dashboard represents the terminal session state.

    Returns
    -------
    str
        Rendered HTML dashboard.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    try:
        raw = conn.execute('SELECT COUNT(*) FROM tweets_raw').fetchone()[0]
        matches = conn.execute('SELECT COUNT(*) FROM tweet_matches').fetchone()[0]
        jobs = pd.read_sql_query("SELECT status, COUNT(*) n FROM x_jobs GROUP BY status", conn)
        job_counts = {r['status']: int(r['n']) for _, r in jobs.iterrows()}
        total_jobs = sum(job_counts.values())
        open_jobs = job_counts.get('pending',0) + job_counts.get('running',0)
        failed_jobs = job_counts.get('failed',0)
        terminal_jobs = total_jobs - open_jobs

        qwp = pd.read_sql_query("""SELECT query_id, project_week,
                 SUM(CASE WHEN status IN ('pending','running') THEN 1 ELSE 0 END) AS open_n,
                 SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) AS failed_n
                 FROM x_jobs GROUP BY query_id, project_week""", conn)
        total_qw = len(X_QUERIES) * len(PROJECT_WEEKS)
        closed_qw = int(((qwp['open_n'] == 0)).sum()) if not qwp.empty else 0
        clean_qw = int(((qwp['open_n'] == 0) & (qwp['failed_n'] == 0)).sum()) if not qwp.empty else 0
        issue_qw = int(((qwp['open_n'] == 0) & (qwp['failed_n'] > 0)).sum()) if not qwp.empty else 0

        workers = pd.read_sql_query('SELECT * FROM worker_heartbeat ORDER BY worker_name', conn)
    finally:
        conn.close()

    elapsed = (time.monotonic() - SESSION_STARTED_MONO) if SESSION_STARTED_MONO else 0
    session_new = int(workers['session_stored'].fillna(0).sum()) if not workers.empty else 0
    rate_h = (session_new / elapsed * 3600) if elapsed > 20 else 0.0
    target_remaining = max(0, RAW_OPERATIONAL_TARGET - raw)
    eta_target = (target_remaining / rate_h * 3600) if rate_h > 0 else None

    qw_pct = _pct(clean_qw, total_qw)
    raw_pct = _pct(raw, RAW_OPERATIONAL_TARGET)
    queue_pct = _pct(terminal_jobs, total_jobs)

    rows = []
    for _, w in workers.iterrows():
        slice_txt = '—'
        if w.get('slice_start'):
            slice_txt = f"{w.get('slice_start','')} → {w.get('slice_end_exclusive','')}"
        hb = w.get('last_heartbeat_utc') or '—'
        rows.append(f"""
        <tr>
          <td><b>{html.escape(str(w.get('account_name') or w.get('worker_name') or ''))}</b><br><small>{html.escape(str(w.get('worker_name') or ''))}</small></td>
          <td>{_badge(w.get('state'))}</td>
          <td><b>{html.escape(str(w.get('query_id') or '—'))}</b> / {html.escape(str(w.get('project_week') or '—'))}<br><small>{html.escape(slice_txt)}</small></td>
          <td>{int(w.get('current_seen') or 0):,}</td>
          <td>{int(w.get('current_stored') or 0):,}</td>
          <td><b>{int(w.get('session_stored') or 0):,}</b></td>
          <td>{int(w.get('session_jobs_done') or 0):,}</td>
          <td>{int(w.get('session_errors') or 0):,}</td>
          <td><small>{html.escape(str(w.get('note') or ''))}</small><br><small>{html.escape(str(hb))}</small></td>
        </tr>""")
    if not rows:
        rows.append('<tr><td colspan="9">Workers have not started yet.</td></tr>')

    final_note = '<div class="final">Final / stopped dashboard</div>' if final else ''
    return f"""
    <style>
      .xwrap{{font-family:Arial,sans-serif;background:#0b1020;color:#e9eefc;padding:16px;border-radius:14px;}}
      .cards{{display:grid;grid-template-columns:repeat(6,minmax(120px,1fr));gap:10px;margin:10px 0 14px;}}
      .card{{background:#151d35;padding:10px;border-radius:10px;border:1px solid #273252;}}
      .big{{font-size:22px;font-weight:700;}} .muted{{color:#9eabc8;font-size:12px;}}
      .bar{{height:10px;background:#252e48;border-radius:8px;overflow:hidden;margin-top:6px;}}
      .fill{{height:100%;background:linear-gradient(90deg,#33c37d,#6f8cff);}}
      table{{width:100%;border-collapse:collapse;background:#11182b;border-radius:10px;overflow:hidden;}}
      th,td{{padding:8px;border-bottom:1px solid #273252;text-align:left;font-size:12px;vertical-align:top;}}
      th{{color:#aeb9d5;background:#171f38;position:sticky;top:0;}}
      .badge{{padding:3px 7px;border-radius:20px;font-weight:700;font-size:10px;white-space:nowrap;}}
      .green{{background:#163d2d;color:#70e0a7}} .amber{{background:#4a3515;color:#ffd37c}}
      .red{{background:#4a1e25;color:#ff909d}} .blue{{background:#1c3151;color:#8fc4ff}} .gray{{background:#333;color:#ccc}}
      .final{{padding:6px 10px;background:#253354;border-radius:8px;display:inline-block;margin-bottom:6px;}}
      @media(max-width:900px){{.cards{{grid-template-columns:repeat(2,1fr);}}}}
    </style>
    <div class="xwrap">
      {final_note}
      <div style="display:flex;justify-content:space-between;align-items:center"><h3 style="margin:0">X Scraper v4.5 · Live Control Panel</h3><span class="muted">Drive write-through: ON · Recovery journal: ON</span></div>
      <div class="cards">
        <div class="card"><div class="muted">Raw unique tweets</div><div class="big">{raw:,}</div><div class="bar"><div class="fill" style="width:{raw_pct:.1f}%"></div></div><div class="muted">Operational target {RAW_OPERATIONAL_TARGET:,} · {raw_pct:.1f}%</div></div>
        <div class="card"><div class="muted">Query×Week clean</div><div class="big">{clean_qw:,}/{total_qw:,}</div><div class="bar"><div class="fill" style="width:{qw_pct:.1f}%"></div></div><div class="muted">{qw_pct:.1f}% · closed with issues {issue_qw}</div></div>
        <div class="card"><div class="muted">Dynamic job queue</div><div class="big">{terminal_jobs:,}/{total_jobs:,}</div><div class="bar"><div class="fill" style="width:{queue_pct:.1f}%"></div></div><div class="muted">open {open_jobs:,} · failed {failed_jobs:,}</div></div>
        <div class="card"><div class="muted">Session new tweets</div><div class="big">{session_new:,}</div><div class="muted">{rate_h:,.0f} new tweets/hour</div></div>
        <div class="card"><div class="muted">Session elapsed</div><div class="big">{_fmt_duration(elapsed)}</div><div class="muted">Matches {matches:,}</div></div>
        <div class="card"><div class="muted">ETA to raw target</div><div class="big">{_fmt_duration(eta_target)}</div><div class="muted">provisional; changes with density/splits</div></div>
      </div>
      <table>
        <thead><tr><th>Account / Worker</th><th>Status</th><th>Current job</th><th>Seen</th><th>New</th><th>Session new</th><th>Jobs</th><th>Errors</th><th>Note / heartbeat</th></tr></thead>
        <tbody>{''.join(rows)}</tbody>
      </table>
      <div class="muted" style="margin-top:10px">pending={job_counts.get('pending',0):,} · running={job_counts.get('running',0):,} · done={job_counts.get('done',0):,} · split={job_counts.get('split',0):,} · failed={job_counts.get('failed',0):,}</div>
    </div>"""

def show_progress_once():
    """Display one snapshot of the persisted collection dashboard.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    display(HTML(progress_dashboard_html()))

def export_raw_csv():
    """Export the main raw X table to CSV.

    Returns
    -------
    Path
        Path to the exported raw CSV file.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    out = EXPORT_DIR / 'x_raw.csv'
    df = pd.read_sql_query('SELECT * FROM tweets_raw ORDER BY created_at_utc, platform_content_id', conn)
    df.to_csv(out, index=False, encoding='utf-8-sig')
    conn.close()
    print(f"Exported {len(df):,} raw rows → {out}")
    return out

def export_subruns_csv():
    """Export detailed adaptive-slice subrun audit records to CSV.

    Returns
    -------
    Path
        Path to the exported subrun CSV file.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    out = EXPORT_DIR / 'x_subruns.csv'
    df = pd.read_sql_query('SELECT * FROM x_subruns ORDER BY project_week, query_id, slice_start, depth', conn)
    df.to_csv(out, index=False, encoding='utf-8-sig')
    conn.close()
    print(f"Exported {len(df):,} subruns → {out}")
    return out

def export_worker_status_csv():
    """Export the latest persisted worker states to CSV.

    Returns
    -------
    Path
        Path to the exported worker-status CSV file.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    out = EXPORT_DIR / 'x_worker_status_last.csv'
    df = pd.read_sql_query('SELECT * FROM worker_heartbeat ORDER BY worker_name', conn)
    df.to_csv(out, index=False, encoding='utf-8-sig')
    conn.close()
    return out

def export_runs_csv():
    # Required audit file: exactly one row per Query × Project Week.
    """Export one audit row per logical Query × Project-Week to CSV.

    Returns
    -------
    Path
        Path to the exported query-week audit CSV file.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    conn = connect_db()
    rows = []
    for q in X_QUERIES:
        for w in PROJECT_WEEKS:
            qid, week = q['query_id'], w['project_week']
            sub = pd.read_sql_query("SELECT * FROM x_subruns WHERE query_id=? AND project_week=?",
                                    conn, params=(qid, week))
            matched = conn.execute("""SELECT COUNT(DISTINCT platform_content_id)
                                      FROM tweet_matches WHERE query_id=? AND project_week=?""",
                                   (qid, week)).fetchone()[0]
            in_window = conn.execute("""SELECT COUNT(DISTINCT m.platform_content_id)
                FROM tweet_matches m JOIN tweets_raw t
                  ON t.platform=m.platform AND t.platform_content_id=m.platform_content_id
                WHERE m.query_id=? AND m.project_week=? AND t.in_window=1""",
                (qid, week)).fetchone()[0]
            dates = conn.execute("""SELECT MIN(t.created_at_utc), MAX(t.created_at_utc)
                FROM tweet_matches m JOIN tweets_raw t
                  ON t.platform=m.platform AND t.platform_content_id=m.platform_content_id
                WHERE m.query_id=? AND m.project_week=? AND t.created_at_utc IS NOT NULL""",
                (qid, week)).fetchone()
            jobs = pd.read_sql_query("SELECT status, depth FROM x_jobs WHERE query_id=? AND project_week=?",
                                     conn, params=(qid, week))
            complete = (not jobs.empty) and not jobs['status'].isin(['pending','running']).any()
            failed_jobs = int((jobs['status'] == 'failed').sum()) if not jobs.empty else 0
            rows.append({
                'collection_run_id': f'XRUN-{qid}-{week}-SUMMARY',
                'platform': 'x',
                'query_id': qid,
                'query_text': q['logical_query'],
                'query_version': PROJECT_CONFIG['query_version'],
                'project_week': week,
                'discovery_route': q['route'],
                'source_id': '',
                'sort_mode': PROJECT_CONFIG['sort_mode'],
                'started_at_utc': sub['started_at_utc'].min() if not sub.empty else '',
                'finished_at_utc': sub['finished_at_utc'].max() if not sub.empty else '',
                'returned_count': matched,
                'stored_count': matched,
                'records_in_window': in_window,
                'sampling_cap': '',
                'records_sampled_out': 0,
                'oldest_record_utc': dates[0] or '',
                'newest_record_utc': dates[1] or '',
                'quota_consumed': '',
                'error_count': (int(sub['error_count'].fillna(0).sum()) if not sub.empty else 0) + failed_jobs,
                'notes': json.dumps({
                    'complete': bool(complete),
                    'subrun_count': int(len(sub)),
                    'max_split_depth': int(jobs['depth'].max()) if not jobs.empty else 0,
                    'failed_jobs': failed_jobs,
                }, ensure_ascii=False)
            })
    df = pd.DataFrame(rows)
    out = EXPORT_DIR / 'x_runs.csv'
    df.to_csv(out, index=False, encoding='utf-8-sig')
    conn.close()
    print(f"Exported {len(df):,} query-week audit rows → {out}")
    return out


def export_excel(download=False):
    """Create the multi-sheet Excel workbook for inspection and audit.

    Parameters
    ----------
    download : bool
        Whether to trigger a Colab browser download after export.

    Returns
    -------
    Path
        Path to the generated Excel workbook.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    out = EXPORT_DIR / 'X_Twitter_Collection.xlsx'
    conn = connect_db()
    raw = pd.read_sql_query('SELECT * FROM tweets_raw ORDER BY created_at_utc, platform_content_id', conn)
    matches = pd.read_sql_query('SELECT * FROM tweet_matches ORDER BY project_week, query_id', conn)
    jobs = pd.read_sql_query('SELECT * FROM x_jobs ORDER BY project_week, query_id, slice_start, depth', conn)
    subruns = pd.read_sql_query('SELECT * FROM x_subruns ORDER BY project_week, query_id, slice_start, depth', conn)
    workers = pd.read_sql_query('SELECT * FROM worker_heartbeat ORDER BY worker_name', conn)
    sessions = pd.read_sql_query('SELECT * FROM collector_sessions ORDER BY started_at_utc', conn)
    conn.close()

    runs_path = EXPORT_DIR / 'x_runs.csv'
    runs = pd.read_csv(runs_path) if runs_path.exists() else pd.DataFrame()

    jc = jobs['status'].value_counts() if not jobs.empty and 'status' in jobs.columns else {}
    summary = pd.DataFrame([
        ['Raw unique tweets', len(raw)],
        ['Tweet-query matches', len(matches)],
        ['Total dynamic jobs', len(jobs)],
        ['Jobs done', int(jc.get('done', 0))],
        ['Jobs split', int(jc.get('split', 0))],
        ['Jobs pending', int(jc.get('pending', 0))],
        ['Jobs running', int(jc.get('running', 0))],
        ['Jobs failed', int(jc.get('failed', 0))],
        ['Rows with username', int(raw['author_username'].notna().sum()) if 'author_username' in raw else 0],
        ['Rows with display name', int(raw['author_display_name'].notna().sum()) if 'author_display_name' in raw else 0],
        ['Rows with tweet URL', int(raw['tweet_url'].notna().sum()) if 'tweet_url' in raw else 0],
    ], columns=['Metric', 'Value'])

    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        raw.to_excel(writer, sheet_name='Raw_Tweets', index=False)
        summary.to_excel(writer, sheet_name='Summary', index=False)
        matches.to_excel(writer, sheet_name='Query_Matches', index=False)
        jobs.to_excel(writer, sheet_name='Jobs', index=False)
        subruns.to_excel(writer, sheet_name='Subruns', index=False)
        workers.to_excel(writer, sheet_name='Worker_Status', index=False)
        sessions.to_excel(writer, sheet_name='Sessions', index=False)
        if not runs.empty:
            runs.to_excel(writer, sheet_name='Query_Week_Audit', index=False)

        wb = writer.book
        for ws in wb.worksheets:
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
            # Sensible widths without scanning the full workbook.
            for col_idx, cells in enumerate(ws.iter_cols(min_row=1, max_row=min(ws.max_row, 250)), start=1):
                max_len = 0
                for cell in cells:
                    if cell.value is not None:
                        max_len = max(max_len, min(len(str(cell.value)), 60))
                ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 60))

        # Make tweet_url directly clickable in the MAIN Raw_Tweets sheet.
        if 'tweet_url' in raw.columns:
            ws = wb['Raw_Tweets']
            url_col = raw.columns.get_loc('tweet_url') + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=url_col)
                if isinstance(cell.value, str) and cell.value.startswith(('https://','http://')):
                    cell.hyperlink = cell.value
                    cell.style = 'Hyperlink'

    print(f"Exported Excel workbook ({len(raw):,} raw rows) → {out}")

    if download:
        try:
            from google.colab import files
            files.download(str(out))
        except Exception as e:
            print(f"Excel created, but automatic browser download was not started: {e}")
    return out


def export_all():
    """Regenerate all supported CSV and Excel exports from the database.

    Returns
    -------
    dict
        Mapping of export names to generated file paths.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    raw_csv = export_raw_csv()
    runs_csv = export_runs_csv()
    subruns_csv = export_subruns_csv()
    worker_csv = export_worker_status_csv()
    excel_file = export_excel(download=False)
    return raw_csv, runs_csv, subruns_csv, worker_csv, excel_file

# -----------------------------------------------------------------------------
# Notebook cell 9
# -----------------------------------------------------------------------------
# QUERY PLAN / PREFLIGHT — run before Main Collection
# This cell is both a human-readable plan and a hard validation gate.
# It reflects Query Registry v3 embedded in X_QUERIES. XQ-H03 is intentionally inactive
# until Persian hashtags are discovered and formally added/backfilled per the Registry.

EXPECTED_ACTIVE_X_QUERY_IDS = {
    'XQ-001','XQ-002','XQ-003','XQ-004','XQ-005','XQ-006','XQ-007','XQ-008','XQ-009',
    'XQ-010','XQ-011','XQ-012','XQ-013','XQ-014','XQ-015','XQ-016','XQ-017','XQ-018',
    'XQ-019','XQ-020','XQ-021','XQ-022','XQ-023','XQ-024','XQ-H01','XQ-H02'
}
INACTIVE_REGISTRY_QUERY_IDS = {'XQ-H03'}


def validate_query_plan(strict=True):
    """Validate the active query registry and project-week plan before collection.

    Parameters
    ----------
    strict : bool
        Whether validation errors should raise an exception.

    Returns
    -------
    dict
        Validation summary containing errors and warnings.

    Raises
    ------
    ValueError
        If strict validation is enabled and the query plan is inconsistent.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if os.getenv("SCRAPER_CUSTOM_TOPIC"):
        return {
                'ok': True,
                'errors': [],
                'warnings': [],
                'query_count': len(X_QUERIES),
                'week_count': len(PROJECT_WEEKS),
                'initial_jobs': len(X_QUERIES) * len(PROJECT_WEEKS),
            }
    errors = []
    warnings = []

    ids = [q.get('query_id') for q in X_QUERIES]
    id_set = set(ids)
    duplicate_ids = sorted({qid for qid in ids if ids.count(qid) > 1})

    if duplicate_ids:
        errors.append(f'Duplicate query IDs: {duplicate_ids}')
    missing = sorted(EXPECTED_ACTIVE_X_QUERY_IDS - id_set)
    extra = sorted(id_set - EXPECTED_ACTIVE_X_QUERY_IDS)
    if missing:
        errors.append(f'Missing active Query Registry IDs: {missing}')
    if extra:
        errors.append(f'Unexpected active query IDs: {extra}')
    if len(X_QUERIES) != 26:
        errors.append(f'Expected 26 active X queries, found {len(X_QUERIES)}')

    if len(PROJECT_WEEKS) != 21:
        errors.append(f'Expected 21 project weeks, found {len(PROJECT_WEEKS)}')
    else:
        if PROJECT_WEEKS[0]['project_week'] != 'W01' or PROJECT_WEEKS[-1]['project_week'] != 'W21':
            errors.append('Project week labels must run W01 → W21')
        if PROJECT_WEEKS[0]['start'].strftime('%Y-%m-%d') != '2026-02-28':
            errors.append('W01 must start 2026-02-28 UTC')
        if PROJECT_WEEKS[-1]['start'].strftime('%Y-%m-%d') != '2026-07-18':
            errors.append('W21 must start 2026-07-18 UTC')
        if PROJECT_WEEKS[-1]['end_exclusive'].strftime('%Y-%m-%d') != '2026-07-23':
            errors.append('W21 end-exclusive must be 2026-07-23 UTC')
        if not PROJECT_WEEKS[-1]['is_partial_week']:
            errors.append('W21 must be marked partial')

    if PROJECT_CONFIG.get('query_version') != 'v3.0':
        errors.append(f"Query version must be v3.0, found {PROJECT_CONFIG.get('query_version')}")
    if PROJECT_CONFIG.get('sort_mode') != 'live':
        errors.append(f"X collection must use live/new sort, found {PROJECT_CONFIG.get('sort_mode')}")

    for q in X_QUERIES:
        qid = q.get('query_id','?')
        if q.get('lang') not in {'en','fa'}:
            errors.append(f'{qid}: invalid lang={q.get("lang")}')
        if q.get('route') not in {'query_search','hashtag'}:
            errors.append(f'{qid}: invalid route={q.get("route")}')
        if not str(q.get('logical_query') or '').strip():
            errors.append(f'{qid}: empty logical_query')
        if q.get('risk') == 'high' and not str(q.get('entity_anchor') or '').strip():
            errors.append(f'{qid}: high-risk query missing entity_anchor')

    if 'XQ-H03' not in INACTIVE_REGISTRY_QUERY_IDS:
        warnings.append('XQ-H03 inactive-state marker is missing')

    result = {
        'ok': not errors,
        'errors': errors,
        'warnings': warnings,
        'query_count': len(X_QUERIES),
        'week_count': len(PROJECT_WEEKS),
        'initial_jobs': len(X_QUERIES) * len(PROJECT_WEEKS),
    }
    if strict and errors:
        raise RuntimeError('QUERY PLAN PREFLIGHT FAILED:\n- ' + '\n- '.join(errors))
    return result


def query_plan_frames():
    """Build data frames summarizing the query registry and project weeks.

    Returns
    -------
    tuple[pandas.DataFrame, pandas.DataFrame]
        Query-registry and project-week data frames.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    detail = pd.DataFrame(X_QUERIES).copy()
    detail['weeks'] = len(PROJECT_WEEKS)
    detail['initial_jobs'] = detail['weeks']
    detail = detail[['query_id','family','lang','risk','route','weeks','initial_jobs','logical_query']]

    topic_rows = []
    for family, grp in detail.groupby('family', sort=False):
        en_ids = grp.loc[grp['lang']=='en','query_id'].tolist()
        fa_ids = grp.loc[grp['lang']=='fa','query_id'].tolist()
        topic_rows.append({
            'Topic family': family,
            'EN': '✓' if en_ids else '—',
            'FA': '✓' if fa_ids else '—',
            'EN Query IDs': ', '.join(en_ids) if en_ids else '—',
            'FA Query IDs': ', '.join(fa_ids) if fa_ids else '—',
            'Queries': len(grp),
            'Weeks/query': len(PROJECT_WEEKS),
            'Initial jobs': len(grp) * len(PROJECT_WEEKS),
        })
    topics = pd.DataFrame(topic_rows)
    return topics, detail


def show_query_plan(show_details=True):
    """Display the human-readable query plan and validation summary.

    Parameters
    ----------
    show_details : bool
        Whether to display detailed query/week tables.

    Returns
    -------
    dict
        Validation summary displayed to the user.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    result = validate_query_plan(strict=False)
    topics, detail = query_plan_frames()
    status = 'PASS' if result['ok'] else 'FAIL'
    status_color = '#163d2d' if result['ok'] else '#4a1e25'
    status_text = '#70e0a7' if result['ok'] else '#ff909d'

    en_count = int((detail['lang']=='en').sum())
    fa_count = int((detail['lang']=='fa').sum())
    high_ids = ', '.join(detail.loc[detail['risk']=='high','query_id'].tolist()) or '—'

    summary_html = f"""
    <div style='font-family:Arial,sans-serif;background:#0b1020;color:#e9eefc;padding:16px;border-radius:14px;margin:8px 0 14px'>
      <div style='display:flex;justify-content:space-between;align-items:center;gap:12px'>
        <h3 style='margin:0'>X Query Plan · Preflight</h3>
        <span style='background:{status_color};color:{status_text};padding:5px 11px;border-radius:20px;font-weight:700'>{status}</span>
      </div>
      <div style='display:grid;grid-template-columns:repeat(6,minmax(110px,1fr));gap:8px;margin-top:12px'>
        <div style='background:#151d35;padding:10px;border-radius:9px'><small>Active queries</small><div style='font-size:22px;font-weight:700'>{result['query_count']}</div></div>
        <div style='background:#151d35;padding:10px;border-radius:9px'><small>English</small><div style='font-size:22px;font-weight:700'>{en_count}</div></div>
        <div style='background:#151d35;padding:10px;border-radius:9px'><small>Persian</small><div style='font-size:22px;font-weight:700'>{fa_count}</div></div>
        <div style='background:#151d35;padding:10px;border-radius:9px'><small>Project weeks</small><div style='font-size:22px;font-weight:700'>{result['week_count']}</div></div>
        <div style='background:#151d35;padding:10px;border-radius:9px'><small>Initial Query×Week jobs</small><div style='font-size:22px;font-weight:700'>{result['initial_jobs']}</div></div>
        <div style='background:#151d35;padding:10px;border-radius:9px'><small>Registry version</small><div style='font-size:22px;font-weight:700'>{html.escape(PROJECT_CONFIG['query_version'])}</div></div>
      </div>
      <div style='margin-top:10px;color:#aeb9d5;font-size:12px'>
        Window: <b>{html.escape(PROJECT_CONFIG['project_start'])} → {html.escape(PROJECT_CONFIG['project_end'])} UTC</b> · W21 partial: <b>2026-07-18 → {html.escape(PROJECT_CONFIG['project_end'])}</b> · sort: <b>{html.escape(PROJECT_CONFIG['sort_mode'])}</b><br>
        High-risk anchored queries: <b>{html.escape(high_ids)}</b> · XQ-H03: <b>inactive pending Persian hashtag discovery + formal backfill</b>
      </div>
    </div>"""
    display(HTML(summary_html))
    display(HTML('<h4>Topic coverage</h4>'))
    display(topics)
    if show_details:
        display(HTML('<h4>Detailed active Query Registry plan</h4>'))
        display(detail)

    if result['errors']:
        display(HTML('<div style="color:#ff909d"><b>Errors:</b><br>' + '<br>'.join(html.escape(x) for x in result['errors']) + '</div>'))
    if result['warnings']:
        display(HTML('<div style="color:#ffd37c"><b>Warnings:</b><br>' + '<br>'.join(html.escape(x) for x in result['warnings']) + '</div>'))
    return result


# Show the plan now. SAFE START runs strict validation again before collection.
QUERY_PLAN_PREFLIGHT = show_query_plan(show_details=True)

# -----------------------------------------------------------------------------
# Notebook cell 12
# -----------------------------------------------------------------------------
# OPTIONAL — LEGACY USERNAME / DISPLAY-NAME BACKFILL (run only while collection is stopped)
# Existing v4.4 rows already receive tweet_url automatically from platform_content_id.
# This utility revisits only rows missing author_username/display name.
# It uses ONE selected existing account/session and stops rather than bypassing login,
# verification/CAPTCHA, or temporary limitation pages.

def backfill_legacy_author_fields(account_index=0, limit=None):
    """Backfill missing legacy author identity fields by revisiting known post IDs.

    Parameters
    ----------
    account_index : int
        Zero-based runtime account index used for backfill.
    limit : int or None
        Maximum number of legacy records to revisit; ``None`` means all.

    Returns
    -------
    dict
        Counts and stop reason for the legacy backfill operation.

    Raises
    ------
    IndexError
        If ``account_index`` does not reference a loaded runtime account.
    SessionInvalid
        If the selected X session is no longer valid.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if account_index < 0 or account_index >= len(ACCOUNTS):
        raise IndexError(f"account_index must be 0..{len(ACCOUNTS)-1}")

    account = ACCOUNTS[account_index]
    account_name = account.get('name', f'Account {account_index+1}')
    conn = connect_db()

    sql = """
        SELECT platform_content_id, tweet_url
        FROM tweets_raw
        WHERE platform='x'
          AND (author_username IS NULL OR TRIM(author_username)=''
               OR author_display_name IS NULL OR TRIM(author_display_name)='')
        ORDER BY created_at_utc, platform_content_id
    """
    rows = conn.execute(sql).fetchall()
    if limit is not None:
        rows = rows[:int(limit)]

    if not rows:
        conn.close()
        print("No legacy author fields need backfill.")
        return {'requested': 0, 'updated': 0, 'stopped_reason': None}

    driver = None
    updated = 0
    stopped_reason = None

    try:
        driver = setup_driver(account, 'backfill')
        if not inject_cookies(driver, account.get('cookies')):
            raise SessionInvalid(f'Login/session validation failed for {account_name}')

        print(f"Backfill using {account_name}: {len(rows):,} legacy row(s)")
        for i, row in enumerate(rows, start=1):
            pid = str(row['platform_content_id'])
            url = row['tweet_url'] or f'https://x.com/i/web/status/{pid}'

            if not safe_get(driver, url, ready_timeout=BACKFILL_READY_TIMEOUT):
                print(f"[{i}/{len(rows)}] navigation failed for {pid}; skipping.")
                time.sleep(random.uniform(BACKFILL_MIN_DELAY, BACKFILL_MAX_DELAY))
                continue

            time.sleep(random.uniform(1.5, 2.8))
            state, note = detect_access_state(driver)
            if state in {'SESSION_INVALID', 'VERIFICATION_REQUIRED', 'TEMP_LIMIT'}:
                stopped_reason = f'{state}: {note}'
                print("Backfill stopped:", stopped_reason)
                break

            try:
                articles = driver.find_elements(By.CSS_SELECTOR, 'article[data-testid="tweet"]')
            except Exception:
                articles = []

            found = False
            for article in articles:
                try:
                    # Find the card matching the requested tweet ID.
                    status_link = None
                    for a in article.find_elements(By.CSS_SELECTOR, 'a[href*="/status/"]'):
                        href = a.get_attribute('href') or ''
                        if extract_status_id(href) == pid:
                            status_link = href
                            break
                    if not status_link:
                        continue

                    handle, display_name = extract_author_identity(article, status_link, pid)
                    direct_url = canonical_tweet_url(handle, pid, status_link)

                    with DB_LOCK:
                        conn.execute("""
                            UPDATE tweets_raw
                            SET author_username=COALESCE(NULLIF(?, ''), author_username),
                                author_display_name=COALESCE(NULLIF(?, ''), author_display_name),
                                tweet_url=COALESCE(NULLIF(?, ''), tweet_url)
                            WHERE platform='x' AND platform_content_id=?
                        """, (handle, display_name, direct_url, pid))
                        conn.commit()
                    updated += 1
                    found = True
                    break
                except Exception:
                    continue

            print(f"[{i}/{len(rows)}] {pid} → {'updated' if found else 'not found'}")
            time.sleep(random.uniform(BACKFILL_MIN_DELAY, BACKFILL_MAX_DELAY))

    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass
        conn.close()

    print(f"Backfill finished: updated={updated:,} / requested={len(rows):,}")
    if stopped_reason:
        print("Stopped reason:", stopped_reason)
    print("Regenerate Excel with: export_excel(download=True)")
    return {'requested': len(rows), 'updated': updated, 'stopped_reason': stopped_reason}

# Safe first test:
# backfill_legacy_author_fields(account_index=1, limit=20)
# Then, if the session stays healthy, continue in small batches rather than forcing a long run:
# backfill_legacy_author_fields(account_index=1, limit=100)


def main():
    """Run query-plan validation, recovery, and the resumable X collection workflow.

    Returns
    -------
    None
        The function runs the collection workflow for its side effects.

    Notes
    -----
    This function is part of the resumable X collection pipeline and preserves
    project provenance and recovery behavior unless stated otherwise.
    """
    if not ACCOUNTS:
        raise RuntimeError(
            "No runtime X account configuration found. "
            f"Set {ACCOUNTS_ENV_VAR} before starting collection."
        )
    validate_query_plan(strict=True)
    recover_from_journals()
    job_stats()
    run_collection(max_workers=MAX_WORKERS)
    job_stats()


if __name__ == '__main__':
    main()
